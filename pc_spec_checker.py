#!/usr/bin/env python3
"""
PC スペックチェッカー
モダンなダークテーマのGUIでPCのスペックを表示するツール
"""

import tkinter as tk
from tkinter import ttk, font, filedialog, messagebox
import platform
import subprocess
import threading
import time
import os
import re
import csv
from datetime import datetime

try:
    import psutil
    HAS_PSUTIL = True
except ImportError:
    HAS_PSUTIL = False

# ── カラーパレット ──────────────────────────────
BG         = "#0d0f14"
PANEL      = "#141820"
BORDER     = "#1e2433"
ACCENT     = "#00d4ff"
ACCENT2    = "#7b4fff"
TEXT_MAIN  = "#e8edf5"
TEXT_SUB   = "#6b7a99"
TEXT_VAL   = "#c0cce6"
GREEN      = "#00e87a"
YELLOW     = "#ffd166"
RED        = "#ff4d6d"
ORANGE     = "#ff8c42"


def run_command(cmd):
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=10,
            shell=(platform.system() == "Windows")
        )
        return result.stdout.strip()
    except Exception:
        return ""


def get_cpu_info():
    info = {}
    sys = platform.system()

    info["名前"] = platform.processor() or "不明"

    if sys == "Darwin":
        brand = run_command(["sysctl", "-n", "machdep.cpu.brand_string"])
        if brand:
            info["名前"] = brand
        cores_physical = run_command(["sysctl", "-n", "hw.physicalcpu"])
        cores_logical  = run_command(["sysctl", "-n", "hw.logicalcpu"])
        if cores_physical:
            info["物理コア数"] = cores_physical
        if cores_logical:
            info["論理コア数"] = cores_logical
        freq = run_command(["sysctl", "-n", "hw.cpufrequency"])
        if freq and freq.isdigit():
            info["基本クロック"] = f"{int(freq) / 1e9:.2f} GHz"

    elif sys == "Linux":
        cpuinfo = run_command(["cat", "/proc/cpuinfo"])
        for line in cpuinfo.splitlines():
            if "model name" in line:
                info["名前"] = line.split(":")[1].strip()
                break
        physical = run_command(
            ["sh", "-c", "grep '^physical id' /proc/cpuinfo | sort -u | wc -l"]
        )
        if physical:
            info["物理コア数"] = physical
        logical = run_command(
            ["sh", "-c", "grep '^processor' /proc/cpuinfo | wc -l"]
        )
        if logical:
            info["論理コア数"] = logical

    elif sys == "Windows":
        name = run_command(
            "wmic cpu get Name /value"
        )
        for line in name.splitlines():
            if "Name=" in line:
                info["名前"] = line.split("=")[1].strip()
        cores = run_command("wmic cpu get NumberOfCores /value")
        for line in cores.splitlines():
            if "NumberOfCores=" in line:
                info["物理コア数"] = line.split("=")[1].strip()
        threads = run_command("wmic cpu get NumberOfLogicalProcessors /value")
        for line in threads.splitlines():
            if "NumberOfLogicalProcessors=" in line:
                info["論理コア数"] = line.split("=")[1].strip()
        speed = run_command("wmic cpu get MaxClockSpeed /value")
        for line in speed.splitlines():
            if "MaxClockSpeed=" in line:
                mhz = line.split("=")[1].strip()
                if mhz.isdigit():
                    info["最大クロック"] = f"{int(mhz)/1000:.2f} GHz"

    if HAS_PSUTIL:
        cpu = psutil.cpu_freq()
        if cpu:
            info["現在クロック"] = f"{cpu.current:.0f} MHz"
        info["物理コア数"] = str(psutil.cpu_count(logical=False))
        info["論理コア数"] = str(psutil.cpu_count(logical=True))

    return info


def get_memory_info():
    info = {}
    sys = platform.system()

    if HAS_PSUTIL:
        mem = psutil.virtual_memory()
        info["合計"] = f"{mem.total / (1024**3):.1f} GB"
        info["使用中"] = f"{mem.used / (1024**3):.1f} GB"
        info["空き"] = f"{mem.available / (1024**3):.1f} GB"
        info["使用率"] = f"{mem.percent:.1f} %"
        swap = psutil.swap_memory()
        info["スワップ合計"] = f"{swap.total / (1024**3):.1f} GB"
        return info

    if sys == "Darwin":
        out = run_command(["sysctl", "-n", "hw.memsize"])
        if out.isdigit():
            info["合計"] = f"{int(out) / (1024**3):.1f} GB"
    elif sys == "Linux":
        out = run_command(["sh", "-c", "grep MemTotal /proc/meminfo"])
        m = re.search(r"(\d+)", out)
        if m:
            info["合計"] = f"{int(m.group(1)) / (1024**2):.1f} GB"
    elif sys == "Windows":
        out = run_command("wmic computersystem get TotalPhysicalMemory /value")
        for line in out.splitlines():
            if "TotalPhysicalMemory=" in line:
                v = line.split("=")[1].strip()
                if v.isdigit():
                    info["合計"] = f"{int(v) / (1024**3):.1f} GB"
    return info


def get_disk_info():
    disks = []
    if HAS_PSUTIL:
        for part in psutil.disk_partitions():
            try:
                usage = psutil.disk_usage(part.mountpoint)
                disks.append({
                    "マウント":   part.mountpoint,
                    "デバイス":   part.device,
                    "ファイルシステム": part.fstype,
                    "合計":      f"{usage.total / (1024**3):.1f} GB",
                    "使用中":    f"{usage.used  / (1024**3):.1f} GB",
                    "空き":      f"{usage.free  / (1024**3):.1f} GB",
                    "使用率":    f"{usage.percent:.1f} %",
                })
            except PermissionError:
                pass
        return disks

    sys = platform.system()
    if sys in ("Darwin", "Linux"):
        out = run_command(["df", "-h"])
        for line in out.splitlines()[1:]:
            parts = line.split()
            if len(parts) >= 6:
                disks.append({
                    "デバイス": parts[0],
                    "合計":    parts[1],
                    "使用中":  parts[2],
                    "空き":    parts[3],
                    "使用率":  parts[4],
                    "マウント": parts[5],
                })
    elif sys == "Windows":
        out = run_command("wmic logicaldisk get Caption,Size,FreeSpace,FileSystem /value")
        disk = {}
        for line in out.splitlines():
            if "=" in line:
                k, v = line.split("=", 1)
                disk[k.strip()] = v.strip()
            elif not line.strip() and disk:
                total = int(disk.get("Size", 0) or 0)
                free  = int(disk.get("FreeSpace", 0) or 0)
                used  = total - free
                disks.append({
                    "デバイス":   disk.get("Caption", ""),
                    "ファイルシステム": disk.get("FileSystem", ""),
                    "合計":      f"{total / (1024**3):.1f} GB" if total else "不明",
                    "使用中":    f"{used  / (1024**3):.1f} GB" if total else "不明",
                    "空き":      f"{free  / (1024**3):.1f} GB" if total else "不明",
                    "使用率":    f"{used/total*100:.1f} %" if total else "不明",
                })
                disk = {}
    return disks


def _ping(ip: str, timeout_ms: int = 300) -> bool:
    """単一IPへのping（Windows/macOS/Linux対応）"""
    sys_name = platform.system()
    timeout_s = max(1, timeout_ms // 1000) if timeout_ms >= 1000 else 1
    try:
        if sys_name == "Windows":
            cmd = ["ping", "-n", "1", "-w", str(timeout_ms), ip]
        else:
            cmd = ["ping", "-c", "1", "-W", str(timeout_s), ip]
        r = subprocess.run(cmd, capture_output=True, timeout=timeout_s + 2)
        return r.returncode == 0
    except Exception:
        return False


def _get_arp_table() -> dict:
    """
    ARPキャッシュから {IP: (MAC, vendor)} を返す。
    vendorはOUIの先頭3オクテットから簡易判定。
    """
    table = {}
    sys_name = platform.system()
    try:
        if sys_name == "Windows":
            r = subprocess.run(["arp", "-a"], capture_output=True,
                               text=True, timeout=10, shell=True)
            for line in r.stdout.splitlines():
                # "  192.168.1.1          aa-bb-cc-dd-ee-ff     動的"
                parts = line.split()
                if len(parts) >= 2:
                    ip_part  = parts[0]
                    mac_part = parts[1]
                    if re.match(r"\d+\.\d+\.\d+\.\d+", ip_part) and                        re.match(r"([0-9a-f]{2}[-:]){5}[0-9a-f]{2}", mac_part, re.I):
                        mac = mac_part.replace("-", ":").upper()
                        table[ip_part] = (mac, _oui_vendor(mac))
        else:
            r = subprocess.run(["arp", "-n"], capture_output=True,
                               text=True, timeout=10)
            for line in r.stdout.splitlines():
                parts = line.split()
                if len(parts) >= 3:
                    ip_part  = parts[0]
                    mac_part = parts[2]
                    if re.match(r"\d+\.\d+\.\d+\.\d+", ip_part) and                        re.match(r"([0-9a-f]{2}:){5}[0-9a-f]{2}", mac_part, re.I):
                        mac = mac_part.upper()
                        table[ip_part] = (mac, _oui_vendor(mac))
    except Exception:
        pass
    return table


# ── OUI データベース ───────────────────────────────────────────────────
# IEEEの公式OUIファイル（oui.txt）をダウンロード・キャッシュして使う。
# ネットワーク不通時や初回起動時はフォールバックテーブルで補完。

_OUI_DB: dict = {}
_OUI_DB_LOADED = False

_OUI_CACHE_PATH = os.path.join(
    os.path.expanduser("~"), ".cache", "pc_spec_checker", "oui.csv"
)

_OUI_FALLBACK = {
    "00:50:56": "VMware",      "00:0C:29": "VMware",
    "08:00:27": "VirtualBox",  "52:54:00": "QEMU/KVM",
    "00:1C:42": "Parallels",   "B8:27:EB": "Raspberry Pi Foundation",
    "DC:A6:32": "Raspberry Pi Foundation", "E4:5F:01": "Raspberry Pi Trading",
    "00:1A:11": "Google",      "54:60:09": "Google",
    "00:23:AE": "Dell",        "14:18:77": "Dell",
    "F8:DB:88": "Dell",        "00:26:B9": "Dell",
    "3C:A9:F4": "Intel",       "00:1E:67": "Intel",
    "8C:8D:28": "Intel",       "00:E0:4C": "Realtek",
    "00:11:32": "Synology",    "00:08:9B": "I-O Data",
    "00:A0:DE": "I-O Data",    "00:10:18": "Buffalo",
    "00:1D:73": "Buffalo",     "7C:DD:90": "Buffalo",
    "AC:DE:48": "Apple",       "28:CD:C1": "Apple",
    "3C:22:FB": "Apple",       "F0:18:98": "Apple",
}

_OUI_TXT_URL = "https://standards-oui.ieee.org/oui/oui.txt"


def _load_oui_db(on_status=None):
    """OUIデータベースをロード。キャッシュ優先、なければ複数URLからダウンロード試行。"""
    global _OUI_DB, _OUI_DB_LOADED

    def notify(msg):
        if on_status:
            try:
                on_status(msg)
            except Exception:
                pass

    # ロード済み & キャッシュ存在 → スキップ
    cache_exists = os.path.isfile(_OUI_CACHE_PATH)
    if _OUI_DB_LOADED and cache_exists:
        notify("OUI DB: 読込済 ({:,} 件)".format(len(_OUI_DB)))
        return
    # ロード済みだがキャッシュ削除済み → リセット
    if _OUI_DB_LOADED and not cache_exists:
        notify("OUI DB: キャッシュが見つかりません。再ダウンロードします...")
        _OUI_DB.clear()
        _OUI_DB_LOADED = False

    # キャッシュが存在すれば読み込む
    if os.path.isfile(_OUI_CACHE_PATH):
        notify("OUI DB: キャッシュを読み込み中...")
        try:
            with open(_OUI_CACHE_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    parts = line.split(",", 1)
                    if len(parts) == 2:
                        _OUI_DB[parts[0].upper()] = parts[1]
            if _OUI_DB:
                _OUI_DB_LOADED = True
                notify("OUI DB: 読込完了 ({:,} 件)".format(len(_OUI_DB)))
                return
        except Exception as ex:
            notify("OUI DB: キャッシュ読込失敗 ({}) → 再ダウンロード".format(ex))
            _OUI_DB.clear()

    # ダウンロード候補URL（順に試す）
    import urllib.request, urllib.error
    HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/plain,*/*",
    }
    CANDIDATE_URLS = [
        ("IEEE (oui.txt)",   "https://standards-oui.ieee.org/oui/oui.txt",   "ieee"),
        ("IEEE (oui.csv)",   "https://standards-oui.ieee.org/oui/oui.csv",   "csv"),
        ("linuxnet.ca",      "https://linuxnet.ca/ieee/oui.txt",              "ieee"),
        ("Wireshark GitHub", "https://github.com/wireshark/wireshark/raw/master/manuf", "manuf"),
        ("arp-scan GitHub",  "https://raw.githubusercontent.com/royhills/arp-scan/master/ieee-oui.txt", "ieee"),
    ]

    # IEEEのoui.txt形式パーサー
    ieee_pat = re.compile(
        "^([0-9A-F]{2}-[0-9A-F]{2}-[0-9A-F]{2})"
        "[	 ]+[(]hex[)][	 ]+(.+)$"
    )
    # IEEE CSV形式パーサー (Registry,Assignment,Organization Name,...)
    csv_pat  = re.compile(
        r"^MA-[LMH]+,([0-9A-F]{6}),\"?([^\"]+)\"?"
    )
    # Wireshark manuf形式パーサー (AA:BB:CC  Vendor  Full name)
    manuf_pat = re.compile(
        r"^([0-9A-Fa-f]{2}:[0-9A-Fa-f]{2}:[0-9A-Fa-f]{2})	([^	]+)"
    )

    def parse_ieee(text):
        db = {}
        for line in text.splitlines():
            m = ieee_pat.match(line.strip())
            if m:
                oui = m.group(1).replace("-", ":").upper()
                db[oui] = m.group(2).strip()
        return db

    def parse_csv(text):
        db = {}
        for line in text.splitlines():
            m = csv_pat.match(line.strip())
            if m:
                raw = m.group(1).upper()
                oui = raw[0:2] + ":" + raw[2:4] + ":" + raw[4:6]
                db[oui] = m.group(2).strip()
        return db

    def parse_manuf(text):
        db = {}
        for line in text.splitlines():
            if line.startswith("#") or not line.strip():
                continue
            m = manuf_pat.match(line)
            if m:
                oui = m.group(1).upper()
                db[oui] = m.group(2).strip()
        return db

    parsers = {"ieee": parse_ieee, "csv": parse_csv, "manuf": parse_manuf}

    os.makedirs(os.path.dirname(_OUI_CACHE_PATH), exist_ok=True)
    tmp = _OUI_CACHE_PATH + ".tmp"

    for label, url, fmt in CANDIDATE_URLS:
        notify("OUI DB: {} からダウンロード中...".format(label))
        try:
            downloaded = [0]
            def reporthook(bn, bs, ts, _u=url, _l=label):
                downloaded[0] = bn * bs
                if ts > 0:
                    pct = min(100, int(downloaded[0] / ts * 100))
                    notify("OUI DB: {} {}%".format(_l, pct))
                else:
                    notify("OUI DB: {} {}KB".format(_l, downloaded[0]//1024))

            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=30) as resp:
                total = int(resp.headers.get("Content-Length", 0))
                with open(tmp, "wb") as tf:
                    block = 8192
                    bn = 0
                    while True:
                        chunk = resp.read(block)
                        if not chunk:
                            break
                        tf.write(chunk)
                        bn += 1
                        reporthook(bn, block, total)

            notify("OUI DB: 解析中 ({})...".format(label))
            with open(tmp, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
            db = parsers[fmt](text)

            if len(db) < 100:
                notify("OUI DB: {} はデータ不足 ({} 件) → 次のURLを試します".format(label, len(db)))
                continue

            notify("OUI DB: キャッシュに保存中...")
            with open(_OUI_CACHE_PATH, "w", encoding="utf-8") as f:
                for k, v in db.items():
                    f.write(k + "," + v + "\n")
            if os.path.exists(tmp):
                os.remove(tmp)

            _OUI_DB.update(db)
            _OUI_DB_LOADED = True
            notify("OUI DB: 完了 ({:,} 件) [{}]".format(len(_OUI_DB), label))
            return

        except urllib.error.HTTPError as e:
            notify("OUI DB: {} HTTP {} → 次のURLを試します".format(label, e.code))
        except Exception as ex:
            notify("OUI DB: {} 失敗 ({}) → 次のURLを試します".format(label, ex))

    # 全URL失敗 → フォールバック
    if os.path.exists(tmp):
        try:
            os.remove(tmp)
        except Exception:
            pass
    _OUI_DB.update(_OUI_FALLBACK)
    _OUI_DB_LOADED = True
    cache_dir = os.path.dirname(_OUI_CACHE_PATH)
    notify("OUI DB: 全URL失敗。フォールバック ({} 件)。"
           "手動配置: {} に oui.txt を置いてください".format(len(_OUI_FALLBACK), cache_dir))

def _oui_vendor(mac: str) -> str:
    """MACアドレス(XX:XX:XX:XX:XX:XX)からベンダー名を返す"""
    # 未ロード or キャッシュが消えた場合は再ロード
    if not _OUI_DB_LOADED or not os.path.isfile(_OUI_CACHE_PATH):
        _load_oui_db()
    prefix = mac[:8].upper().replace("-", ":")
    return _OUI_DB.get(prefix, _OUI_FALLBACK.get(prefix, ""))


def _ensure_oui_db_async(on_status=None):
    """バックグラウンドでOUIデータベースを非同期ロードする
    on_status: 状態文字列を受け取るコールバック関数 (省略可)
    """
    import threading as _th
    _th.Thread(target=_load_oui_db, args=(on_status,), daemon=True).start()


def get_printer_info():
    """インストール済みプリンター一覧を返す。戻り値: list of dict"""
    printers = []
    sys_name = platform.system()

    if sys_name == "Windows":
        import tempfile

        STATUS_MAP = {
            "1": "その他", "2": "不明", "3": "待機中", "4": "印刷中",
            "5": "暖機中", "6": "停止中", "7": "オフライン",
        }

        def _parse_printer_lines(text):
            """---PRINTER--- 区切りのテキストをパースして list of dict を返す"""
            result = []
            mapping = {
                "NAME": "プリンター名", "DRIVER": "ドライバー",
                "PORT": "ポート",       "STATUS": "状態",
                "DEFAULT": "既定",      "SHARED": "共有",
                "COMMENT": "コメント",
            }
            cur = {}
            for raw in text.splitlines():
                raw = raw.strip()
                if raw == "---PRINTER---":
                    if cur.get("プリンター名"):
                        result.append(cur)
                    cur = {}
                elif raw.startswith("STATUS:"):
                    val = raw[7:].strip()
                    cur["状態"] = STATUS_MAP.get(val, val) if val else "不明"
                elif ":" in raw:
                    key, val = raw.split(":", 1)
                    val = val.strip()
                    label = mapping.get(key.strip())
                    if label and val:
                        cur[label] = val
            if cur.get("プリンター名"):
                result.append(cur)
            return result

        # ── 方法1: PowerShell ファイル実行 ──────────────────
        ps_script = (
            "$ErrorActionPreference='SilentlyContinue'\n"
            "$list = Get-CimInstance Win32_Printer\n"
            "if(-not $list){$list=Get-WmiObject Win32_Printer}\n"
            "foreach($p in $list){\n"
            "  Write-Output '---PRINTER---'\n"
            "  Write-Output ('NAME:'    + $p.Name)\n"
            "  Write-Output ('DRIVER:'  + $p.DriverName)\n"
            "  Write-Output ('PORT:'    + $p.PortName)\n"
            "  Write-Output ('STATUS:'  + $p.PrinterStatus)\n"
            "  Write-Output ('DEFAULT:' + $(if($p.Default){'はい'}else{'いいえ'}))\n"
            "  Write-Output ('SHARED:'  + $(if($p.Shared) {'はい'}else{'いいえ'}))\n"
            "  if($p.Comment){Write-Output ('COMMENT:' + $p.Comment)}\n"
            "}\n"
        )
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(mode='w', suffix='.ps1',
                                             delete=False, encoding='utf-8') as tf:
                tf.write(ps_script)
                tmp_path = tf.name
            for ps_exe in ["pwsh", "powershell"]:
                try:
                    r = subprocess.run(
                        [ps_exe, "-NoProfile", "-NonInteractive",
                         "-ExecutionPolicy", "Bypass", "-File", tmp_path],
                        capture_output=True, text=True, timeout=30
                    )
                    if r.returncode == 0 and "---PRINTER---" in r.stdout:
                        printers = _parse_printer_lines(r.stdout)
                        break
                except FileNotFoundError:
                    continue
        except Exception:
            pass
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

        # ── 方法2: PowerShell -Command（短いコマンド）────────
        if not printers:
            try:
                cmd = (
                    "Get-CimInstance Win32_Printer | "
                    "ForEach-Object { "
                    "Write-Output '---PRINTER---'; "
                    "Write-Output ('NAME:'+$_.Name); "
                    "Write-Output ('DRIVER:'+$_.DriverName); "
                    "Write-Output ('PORT:'+$_.PortName); "
                    "Write-Output ('STATUS:'+$_.PrinterStatus); "
                    "Write-Output ('DEFAULT:'+$(if($_.Default){'はい'}else{'いいえ'})); "
                    "Write-Output ('SHARED:'+$(if($_.Shared){'はい'}else{'いいえ'})) }"
                )
                for ps_exe in ["pwsh", "powershell"]:
                    try:
                        r = subprocess.run(
                            [ps_exe, "-NoProfile", "-NonInteractive",
                             "-ExecutionPolicy", "Bypass", "-Command", cmd],
                            capture_output=True, text=True, timeout=30
                        )
                        if r.returncode == 0 and "---PRINTER---" in r.stdout:
                            printers = _parse_printer_lines(r.stdout)
                            break
                    except FileNotFoundError:
                        continue
            except Exception:
                pass

        # ── 方法3: wmic フォールバック ────────────────────────
        if not printers:
            try:
                r = subprocess.run(
                    "wmic printer get Name,DriverName,PortName,Default,Shared,PrinterStatus /value",
                    capture_output=True, text=True, timeout=20, shell=True
                )
                cur = {}
                def flush(c):
                    if c.get("プリンター名"):
                        printers.append(c)
                for raw in r.stdout.splitlines():
                    raw = raw.strip()
                    if not raw:
                        flush(cur); cur = {}
                        continue
                    if "=" not in raw:
                        continue
                    k, v = raw.split("=", 1)
                    k, v = k.strip(), v.strip()
                    if not v:
                        continue
                    if k == "Name":
                        cur["プリンター名"] = v
                    elif k == "DriverName":
                        cur["ドライバー"] = v
                    elif k == "PortName":
                        cur["ポート"] = v
                    elif k == "Default":
                        cur["既定"] = "はい" if v.upper() == "TRUE" else "いいえ"
                    elif k == "Shared":
                        cur["共有"] = "はい" if v.upper() == "TRUE" else "いいえ"
                    elif k == "PrinterStatus":
                        cur["状態"] = STATUS_MAP.get(v, v)
                flush(cur)
            except Exception:
                pass

    elif sys_name == "Darwin":
        # lpstat でプリンター一覧
        lp_out = run_command(["lpstat", "-p", "-d"])
        current_name = None
        for line in lp_out.splitlines():
            line = line.strip()
            if line.startswith("printer "):
                parts = line.split()
                current_name = parts[1] if len(parts) > 1 else line
                status = "待機中"
                if "is idle" in line:
                    status = "待機中"
                elif "is busy" in line or "printing" in line:
                    status = "印刷中"
                elif "disabled" in line:
                    status = "無効"
                printers.append({"プリンター名": current_name, "状態": status})
            elif line.startswith("system default destination:"):
                default_name = line.split(":", 1)[1].strip()
                for p in printers:
                    if p["プリンター名"] == default_name:
                        p["既定"] = "はい"

        # lpinfo でドライバー情報を補完
        lpinfo_out = run_command(["lpinfo", "-l", "-v"])
        for p in printers:
            detail_out = run_command(["lpoptions", "-p", p["プリンター名"], "-l"])
            if detail_out:
                p["オプション数"] = f"{len(detail_out.splitlines())} 項目"

    elif sys_name == "Linux":
        # ── 方法1: python3-cups (pycups) ─────────────────────
        try:
            import cups
            conn = cups.Connection()
            dests = conn.getDests()
            default_name = ""
            # デフォルトプリンター名を取得
            try:
                default_name = conn.getDefault() or ""
            except Exception:
                pass
            # プリンター詳細を取得
            printers_detail = {}
            try:
                printers_detail = conn.getPrinters()
            except Exception:
                pass
            for (name, inst), dest in dests.items():
                if not name:
                    continue
                info = printers_detail.get(name, {})
                state_code = info.get("printer-state", 0)
                state_map = {3: "待機中", 4: "印刷中", 5: "停止"}
                status = state_map.get(state_code, "不明")
                p = {
                    "プリンター名": name,
                    "状態": status,
                }
                if name == default_name:
                    p["既定"] = "はい"
                uri = info.get("device-uri", "")
                if uri:
                    p["URI"] = uri
                location = info.get("printer-location", "")
                if location:
                    p["場所"] = location
                driver = info.get("printer-make-and-model", "")
                if driver:
                    p["ドライバー"] = driver
                printers.append(p)
        except ImportError:
            pass
        except Exception:
            pass

        # ── 方法2: lpstat コマンド ───────────────────────────
        if not printers:
            lp_out = run_command(["lpstat", "-p", "-d"])
            if not lp_out:
                # CUPS サービスが停止している場合は起動を試みる（権限があれば）
                run_command(["sh", "-c", "systemctl start cups 2>/dev/null || true"])
                lp_out = run_command(["lpstat", "-p", "-d"])
            for line in lp_out.splitlines():
                line_stripped = line.strip()
                if line_stripped.startswith("printer "):
                    parts = line_stripped.split()
                    name = parts[1] if len(parts) > 1 else ""
                    if not name:
                        continue
                    status = "待機中"
                    if "is idle" in line_stripped:
                        status = "待機中"
                    elif "printing" in line_stripped or "is busy" in line_stripped:
                        status = "印刷中"
                    elif "disabled" in line_stripped or "stopped" in line_stripped:
                        status = "停止"
                    printers.append({"プリンター名": name, "状態": status})
                elif line_stripped.startswith("system default destination:"):
                    default_name = line_stripped.split(":", 1)[1].strip()
                    for p in printers:
                        if p["プリンター名"] == default_name:
                            p["既定"] = "はい"

            # CUPS URI を取得 ("device for <name>: <uri>" 形式に対応)
            if printers:
                uri_all = run_command(["lpstat", "-v"])
                uri_map = {}
                for line in uri_all.splitlines():
                    # "device for PrinterName: uri://..."
                    m = re.match(r"device for ([^:]+):\s*(.+)", line.strip())
                    if m:
                        uri_map[m.group(1).strip()] = m.group(2).strip()
                for p in printers:
                    uri = uri_map.get(p["プリンター名"], "")
                    if uri:
                        p["URI"] = uri

            # ドライバー情報を lpinfo で補完
            for p in printers:
                detail_out = run_command(
                    ["lpoptions", "-p", p["プリンター名"], "-l"]
                )
                if detail_out:
                    p["オプション数"] = f"{len(detail_out.splitlines())} 項目"

        # ── 方法3: /etc/cups/printers.conf を直接読む ─────────
        if not printers:
            conf_path = "/etc/cups/printers.conf"
            try:
                with open(conf_path, "r", encoding="utf-8", errors="replace") as f:
                    conf_text = f.read()
                cur = {}
                for line in conf_text.splitlines():
                    line = line.strip()
                    if line.startswith("<Printer "):
                        cur = {"プリンター名": line[9:].rstrip(">")}
                    elif line == "</Printer>" and cur.get("プリンター名"):
                        printers.append(cur)
                        cur = {}
                    elif line.startswith("DeviceURI "):
                        cur["URI"] = line[10:].strip()
                    elif line.startswith("Info "):
                        cur["情報"] = line[5:].strip()
                    elif line.startswith("MakeModel "):
                        cur["ドライバー"] = line[10:].strip()
                    elif line.startswith("State "):
                        state = line[6:].strip()
                        cur["状態"] = {"Idle": "待機中", "Processing": "印刷中",
                                       "Stopped": "停止"}.get(state, state)
                    elif line == "DefaultPrinter" or line.startswith("DefaultPrinter "):
                        if cur.get("プリンター名"):
                            cur["既定"] = "はい"
            except PermissionError:
                printers.append({"備考": "/etc/cups/printers.conf の読み取り権限がありません (sudo が必要)"})
            except FileNotFoundError:
                pass

    if not printers:
        printers.append({"備考": "プリンターが見つかりませんでした"})

    return printers


def get_installed_apps():
    """インストール済みアプリ一覧を返す。戻り値: list of dict"""
    apps = []
    sys_name = platform.system()

    if sys_name == "Darwin":
        # /Applications 以下の .app を列挙
        app_dirs = ["/Applications", os.path.expanduser("~/Applications")]
        seen = set()
        for base in app_dirs:
            if not os.path.isdir(base):
                continue
            for name in sorted(os.listdir(base)):
                if not name.endswith(".app"):
                    continue
                if name in seen:
                    continue
                seen.add(name)
                display = name[:-4]
                info_plist = os.path.join(base, name, "Contents", "Info.plist")
                version = ""
                try:
                    # plutil で plist をパース
                    out = run_command(["plutil", "-extract", "CFBundleShortVersionString",
                                       "raw", "-o", "-", info_plist])
                    version = out.strip() if out else ""
                except Exception:
                    pass
                apps.append({"アプリ名": display, "バージョン": version,
                             "場所": base})

    elif sys_name == "Linux":
        # dpkg (Debian/Ubuntu 系)
        dpkg_out = run_command(["dpkg-query", "-W",
                                "-f=${Package}\t${Version}\t${Status}\n"])
        if dpkg_out:
            for line in dpkg_out.splitlines():
                parts = line.split("\t")
                if len(parts) >= 3 and "installed" in parts[2]:
                    apps.append({"パッケージ名": parts[0],
                                 "バージョン": parts[1],
                                 "管理": "dpkg"})
        # rpm (RHEL/Fedora/CentOS 系)
        if not apps:
            rpm_out = run_command(["rpm", "-qa", "--queryformat",
                                   "%{NAME}\t%{VERSION}-%{RELEASE}\n"])
            if rpm_out:
                for line in rpm_out.splitlines():
                    parts = line.split("\t")
                    if len(parts) >= 2:
                        apps.append({"パッケージ名": parts[0],
                                     "バージョン": parts[1],
                                     "管理": "rpm"})
        # flatpak
        fp_out = run_command(["flatpak", "list", "--app",
                              "--columns=application,version"])
        if fp_out:
            for line in fp_out.splitlines():
                parts = line.split("\t")
                if parts:
                    apps.append({"パッケージ名": parts[0],
                                 "バージョン": parts[1] if len(parts) > 1 else "",
                                 "管理": "flatpak"})

    elif sys_name == "Windows":
        # PowerShell スクリプトを一時ファイルに書き出して実行
        # （インライン -Command は長いスクリプトで構文エラーになりやすいため）
        import tempfile, sys as _sys
        ps_script = r"""
$paths = @(
    'HKLM:\Software\Microsoft\Windows\CurrentVersion\Uninstall\*',
    'HKLM:\Software\Wow6432Node\Microsoft\Windows\CurrentVersion\Uninstall\*',
    'HKCU:\Software\Microsoft\Windows\CurrentVersion\Uninstall\*'
)
$seen = @{}
foreach ($p in $paths) {
    $items = Get-ItemProperty $p -ErrorAction SilentlyContinue
    foreach ($item in $items) {
        $name = $item.DisplayName
        if (-not $name -or $name -eq '') { continue }
        if ($seen[$name]) { continue }
        $seen[$name] = 1
        $ver = if ($item.DisplayVersion) { $item.DisplayVersion } else { '' }
        $pub = if ($item.Publisher)      { $item.Publisher }      else { '' }
        $dat = if ($item.InstallDate)    { $item.InstallDate }    else { '' }
        # タブ区切りで出力（アプリ名に | が含まれる場合も安全）
        Write-Output ("{0}`t{1}`t{2}`t{3}" -f $name, $ver, $pub, $dat)
    }
}
"""
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(mode='w', suffix='.ps1',
                                             delete=False, encoding='utf-8') as tf:
                tf.write(ps_script)
                tmp_path = tf.name

            result = subprocess.run(
                ["powershell", "-NoProfile", "-NonInteractive",
                 "-ExecutionPolicy", "Bypass", "-File", tmp_path],
                capture_output=True, text=True, timeout=60
            )
            lines = result.stdout.splitlines()
            # 名前でソート
            lines.sort(key=lambda x: x.lower())
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                parts = line.split("\t")
                entry = {"アプリ名": parts[0]}
                if len(parts) > 1 and parts[1].strip():
                    entry["バージョン"] = parts[1].strip()
                if len(parts) > 2 and parts[2].strip():
                    entry["発行元"] = parts[2].strip()
                if len(parts) > 3 and parts[3].strip():
                    raw = parts[3].strip()
                    # YYYYMMDD → YYYY/MM/DD に整形
                    if len(raw) == 8 and raw.isdigit():
                        raw = f"{raw[:4]}/{raw[4:6]}/{raw[6:]}"
                    entry["インストール日"] = raw
                apps.append(entry)
        except Exception:
            pass
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)

    if not apps:
        apps.append({"備考": "アプリ情報を取得できませんでした"})

    return apps


def get_network_info():
    """
    各ネットワークインターフェースの情報を返す。
    戻り値: list of dict  (インターフェースごとに1つの dict)
    """
    import uuid as _uuid
    import socket as _socket

    interfaces = []

    if HAS_PSUTIL:
        addrs = psutil.net_if_addrs()
        stats = psutil.net_if_stats()
        AF_INET  = _socket.AF_INET
        AF_INET6 = _socket.AF_INET6
        # psutil では MAC は AF_LINK (17) または AF_PACKET (17)
        MAC_FAMILIES = {psutil.AF_LINK} if hasattr(psutil, "AF_LINK") else set()
        # fallback: family value == 17 (AF_PACKET on Linux, AF_LINK on macOS/Win)
        MAC_FAMILIES.add(17)

        for iface, addr_list in addrs.items():
            st = stats.get(iface)
            ipv4_obj = next((a for a in addr_list if a.family == AF_INET), None)
            ipv6 = next((a.address.split("%")[0] for a in addr_list if a.family == AF_INET6
                         and not a.address.startswith("fe80")), None)
            mac  = next((a.address for a in addr_list if a.family in MAC_FAMILIES), None)

            ipv4    = ipv4_obj.address if ipv4_obj else None
            netmask = ipv4_obj.netmask if ipv4_obj else None

            # ループバックは除外
            if ipv4 == "127.0.0.1" and not mac:
                continue
            # 何も情報がないインターフェースはスキップ
            if not ipv4 and not mac:
                continue

            entry = {"インターフェース": iface}
            if mac:
                entry["MACアドレス"] = mac.upper()
            if ipv4:
                entry["IPv4アドレス"] = ipv4
            if netmask:
                entry["サブネットマスク"] = netmask
            if ipv6:
                entry["IPv6アドレス"] = ipv6
            if st:
                entry["リンク速度"] = f"{st.speed} Mbps" if st.speed else "不明"
                entry["状態"] = "稼働中" if st.isup else "停止中"
            interfaces.append(entry)
        return interfaces

    # ── psutil なし: 標準ライブラリ + OS コマンドで取得 ──
    sys_name = platform.system()

    # uuid モジュールで主インターフェースの MAC だけでも取得
    raw_mac = _uuid.getnode()
    mac_str = ":".join(f"{(raw_mac >> (8*i)) & 0xff:02X}" for i in reversed(range(6)))

    if sys_name == "Darwin":
        out = run_command(["ifconfig"])
        current_iface = None
        iface_data = {}
        for line in out.splitlines():
            m = re.match(r"^(\S+):", line)
            if m:
                if current_iface and iface_data:
                    interfaces.append(iface_data)
                current_iface = m.group(1)
                iface_data = {"インターフェース": current_iface}
            elif current_iface:
                m_mac  = re.search(r"ether\s+([0-9a-f:]{17})", line)
                m_ipv4 = re.search(r"inet\s+(\d+\.\d+\.\d+\.\d+)", line)
                m_mask = re.search(r"netmask\s+(0x[0-9a-fA-F]+|\d+\.\d+\.\d+\.\d+)", line)
                m_ipv6 = re.search(r"inet6\s+([0-9a-f:]+)(?:%\S+)?", line)
                if m_mac:
                    iface_data["MACアドレス"] = m_mac.group(1).upper()
                if m_ipv4 and m_ipv4.group(1) != "127.0.0.1":
                    iface_data["IPv4アドレス"] = m_ipv4.group(1)
                if m_mask:
                    raw = m_mask.group(1)
                    if raw.startswith("0x"):
                        val = int(raw, 16)
                        mask = ".".join(str((val >> (8*i)) & 0xff) for i in reversed(range(4)))
                    else:
                        mask = raw
                    iface_data["サブネットマスク"] = mask
                if m_ipv6 and not m_ipv6.group(1).startswith("::1"):
                    iface_data.setdefault("IPv6アドレス", m_ipv6.group(1))
        if current_iface and iface_data:
            interfaces.append(iface_data)

    elif sys_name == "Linux":
        out = run_command(["ip", "addr"])
        current_iface = None
        iface_data = {}
        for line in out.splitlines():
            m = re.match(r"^\d+:\s+(\S+):", line)
            if m:
                if current_iface and iface_data:
                    interfaces.append(iface_data)
                current_iface = m.group(1).rstrip("@")
                iface_data = {"インターフェース": current_iface}
            elif current_iface:
                m_mac  = re.search(r"link/ether\s+([0-9a-f:]{17})", line)
                m_ipv4 = re.search(r"inet\s+(\d+\.\d+\.\d+\.\d+)/(\d+)", line)
                m_ipv6 = re.search(r"inet6\s+([0-9a-f:]+)(?:/\d+)?", line)
                if m_mac:
                    iface_data["MACアドレス"] = m_mac.group(1).upper()
                if m_ipv4 and m_ipv4.group(1) != "127.0.0.1":
                    iface_data["IPv4アドレス"] = m_ipv4.group(1)
                    prefix = int(m_ipv4.group(2))
                    bits = (0xFFFFFFFF << (32 - prefix)) & 0xFFFFFFFF
                    mask = ".".join(str((bits >> (8*i)) & 0xff) for i in reversed(range(4)))
                    iface_data["サブネットマスク"] = mask
                if m_ipv6 and not m_ipv6.group(1).startswith("::1") \
                           and not m_ipv6.group(1).startswith("fe80"):
                    iface_data.setdefault("IPv6アドレス", m_ipv6.group(1))
        if current_iface and iface_data:
            interfaces.append(iface_data)

    elif sys_name == "Windows":
        out = run_command("getmac /v /fo list")
        current = {}
        for line in out.splitlines():
            line = line.strip()
            if not line:
                if current:
                    interfaces.append(current)
                    current = {}
                continue
            if "接続名" in line or "Connection Name" in line:
                current["インターフェース"] = line.split(":", 1)[1].strip()
            elif "物理アドレス" in line or "Physical Address" in line:
                current["MACアドレス"] = line.split(":", 1)[1].strip().upper()
            elif "トランスポート名" in line or "Transport Name" in line:
                current["トランスポート"] = line.split(":", 1)[1].strip()
        if current:
            interfaces.append(current)

        # IPv4・サブネットマスクは ipconfig から補完
        ipcfg = run_command("ipconfig")
        iface_name = None
        for line in ipcfg.splitlines():
            m = re.match(r"^(\S.*):$", line)
            if m:
                iface_name = m.group(1).strip()
            m_ip   = re.search(r"IPv4.*?:\s+(\d+\.\d+\.\d+\.\d+)", line)
            m_mask = re.search(r"サブネット マスク.*?:\s+(\d+\.\d+\.\d+\.\d+)|Subnet Mask.*?:\s+(\d+\.\d+\.\d+\.\d+)", line)
            if m_ip and iface_name:
                for entry in interfaces:
                    if iface_name in entry.get("インターフェース", ""):
                        entry["IPv4アドレス"] = m_ip.group(1)
            if m_mask and iface_name:
                mask_val = m_mask.group(1) or m_mask.group(2)
                for entry in interfaces:
                    if iface_name in entry.get("インターフェース", ""):
                        entry["サブネットマスク"] = mask_val

    if not interfaces:
        interfaces.append({"インターフェース": "不明", "MACアドレス": mac_str})

    return interfaces


def get_machine_info():
    """機種名とシリアルナンバーを返す (macOS / Linux / Windows 対応)"""
    info = {}
    sys_name = platform.system()

    if sys_name == "Darwin":
        sp = run_command(["system_profiler", "SPHardwareDataType"])
        for line in sp.splitlines():
            line = line.strip()
            if "Model Name:" in line:
                info["モデル名"] = line.split(":", 1)[1].strip()
            elif "Model Identifier:" in line:
                info["モデル識別子"] = line.split(":", 1)[1].strip()
            elif "Chip:" in line:
                info["チップ"] = line.split(":", 1)[1].strip()
            elif "Serial Number" in line:
                info["シリアルNo."] = line.split(":", 1)[1].strip()
            elif "Hardware UUID" in line:
                info["ハードウェアUUID"] = line.split(":", 1)[1].strip()
        # sysctl フォールバック
        if not info.get("モデル識別子"):
            m = run_command(["sysctl", "-n", "hw.model"])
            if m:
                info["モデル識別子"] = m

    elif sys_name == "Linux":
        # 優先: /sys/class/dmi/id (root不要、多くのディストリで読める)
        import os
        dmi_map = [
            ("sys_vendor",      "メーカー"),
            ("product_name",    "機種名"),
            ("product_version", "製品バージョン"),
            ("product_serial",  "シリアルNo."),
            ("product_uuid",    "UUID"),
            ("board_name",      "ボード名"),
            ("board_vendor",    "ボードメーカー"),
        ]
        for dmi_key, label in dmi_map:
            path = f"/sys/class/dmi/id/{dmi_key}"
            try:
                with open(path, "r") as f:
                    val = f.read().strip()
                skip = {"", "To Be Filled By O.E.M.", "Default string",
                        "None", "N/A", "Not Specified", "Unknown"}
                if val not in skip:
                    info[label] = val
            except Exception:
                pass

        # 次善: dmidecode (root 権限があれば)
        if not info:
            out = run_command(["dmidecode", "-t", "system"])
            for line in out.splitlines():
                line = line.strip()
                for keyword, label in [
                    ("Manufacturer:", "メーカー"),
                    ("Product Name:", "機種名"),
                    ("Serial Number:", "シリアルNo."),
                    ("UUID:", "UUID"),
                ]:
                    if line.startswith(keyword):
                        val = line.split(":", 1)[1].strip()
                        skip = {"Not Specified", "Not Present", "", "None"}
                        if val not in skip:
                            info[label] = val

        # さらに: /proc/cpuinfo の Hardware / Revision (Raspberry Pi 等)
        if not info:
            cpuinfo = run_command(["cat", "/proc/cpuinfo"])
            for line in cpuinfo.splitlines():
                if line.startswith("Hardware"):
                    info["ハードウェア"] = line.split(":", 1)[1].strip()
                elif line.startswith("Revision"):
                    info["リビジョン"] = line.split(":", 1)[1].strip()
                elif line.startswith("Serial"):
                    info["シリアルNo."] = line.split(":", 1)[1].strip()

        if not info:
            info["備考"] = "DMI情報を取得できませんでした (仮想環境または権限不足)"

    elif sys_name == "Windows":
        # PowerShell で一括取得（1回の起動で全情報を取得してタイムアウトを最小化）
        import tempfile as _tf2
        import os as _os2
        ps_script = (
            "$cs   = Get-CimInstance Win32_ComputerSystem -ErrorAction SilentlyContinue\n"
            "$bios = Get-CimInstance Win32_BIOS -ErrorAction SilentlyContinue\n"
            "$csp  = Get-CimInstance Win32_ComputerSystemProduct -ErrorAction SilentlyContinue\n"
            "if(-not $cs){$cs   = Get-WmiObject Win32_ComputerSystem -ErrorAction SilentlyContinue}\n"
            "if(-not $bios){$bios = Get-WmiObject Win32_BIOS -ErrorAction SilentlyContinue}\n"
            "if(-not $csp){$csp  = Get-WmiObject Win32_ComputerSystemProduct -ErrorAction SilentlyContinue}\n"
            "Write-Output ('MANUFACTURER:' + $cs.Manufacturer)\n"
            "Write-Output ('MODEL:'        + $cs.Model)\n"
            "Write-Output ('FAMILY:'       + $cs.SystemFamily)\n"
            "Write-Output ('SERIAL:'       + $bios.SerialNumber)\n"
            "Write-Output ('BIOSVER:'      + $bios.SMBIOSBIOSVersion)\n"
            "Write-Output ('UUID:'         + $csp.UUID)\n"
        )
        skip_vals = {"", "None", "N/A", "Not Available", "To Be Filled By O.E.M.",
                     "Default string", "System Product Name", "System Manufacturer",
                     "System Version", "INVALID", "00000000-0000-0000-0000-000000000000"}
        mapping = {
            "MANUFACTURER": "メーカー",
            "MODEL":        "機種名",
            "FAMILY":       "製品ファミリー",
            "SERIAL":       "シリアルNo.",
            "BIOSVER":      "BIOSバージョン",
            "UUID":         "UUID",
        }
        tmp_path = None
        try:
            with _tf2.NamedTemporaryFile(mode='w', suffix='.ps1',
                                         delete=False, encoding='utf-8') as tf:
                tf.write(ps_script)
                tmp_path = tf.name
            for ps_exe in ["pwsh", "powershell"]:
                try:
                    r = subprocess.run(
                        [ps_exe, "-NoProfile", "-NonInteractive",
                         "-ExecutionPolicy", "Bypass", "-File", tmp_path],
                        capture_output=True, text=True, timeout=20
                    )
                    if r.stdout.strip():
                        for line in r.stdout.splitlines():
                            line = line.strip()
                            if ":" not in line:
                                continue
                            key, val = line.split(":", 1)
                            val = val.strip()
                            label = mapping.get(key.strip())
                            if label and val and val not in skip_vals:
                                info[label] = val
                        break
                except FileNotFoundError:
                    continue
                except subprocess.TimeoutExpired:
                    break
        except Exception:
            pass
        finally:
            if tmp_path and _os2.path.exists(tmp_path):
                try:
                    _os2.remove(tmp_path)
                except Exception:
                    pass

        # フォールバック: winreg で直接読む（PowerShell不要・即時）
        if "機種名" not in info or "シリアルNo." not in info:
            try:
                import winreg
                reg_key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                                         r"HARDWARE\DESCRIPTION\System\BIOS")
                reg_map = {
                    "SystemManufacturer": "メーカー",
                    "SystemProductName":  "機種名",
                    "SystemVersion":      "製品バージョン",
                    "BIOSVersion":        "BIOSバージョン",
                }
                for reg_name, label in reg_map.items():
                    if label in info:
                        continue
                    try:
                        val, _ = winreg.QueryValueEx(reg_key, reg_name)
                        if val and val not in skip_vals:
                            info[label] = val
                    except Exception:
                        pass
                winreg.CloseKey(reg_key)
            except Exception:
                pass

        if not info:
            info["備考"] = "機種情報を取得できませんでした"
    return info

def get_os_info():
    info = {
        "OS":           platform.system(),
        "バージョン":   platform.version(),
        "リリース":     platform.release(),
        "アーキテクチャ": platform.machine(),
        "ホスト名":     platform.node(),
        "Pythonバージョン": platform.python_version(),
    }
    sys = platform.system()
    if sys == "Darwin":
        ver = run_command(["sw_vers", "-productVersion"])
        name = run_command(["sw_vers", "-productName"])
        if ver:
            info["macOSバージョン"] = f"{name} {ver}"
    elif sys == "Linux":
        out = run_command(["cat", "/etc/os-release"])
        for line in out.splitlines():
            if line.startswith("PRETTY_NAME="):
                info["ディストリビューション"] = line.split("=")[1].strip().strip('"')
    return info


def _ps_run(script):
    """PowerShell コマンドを実行して stdout を返す"""
    try:
        r = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", script],
            capture_output=True, text=True, timeout=15
        )
        return r.stdout.strip()
    except Exception:
        return ""


def get_gpu_info():
    """GPU情報を返す。戻り値は list of dict (GPU ごとに 1 dict)"""
    gpus = []
    sys_name = platform.system()

    if sys_name == "Darwin":
        # system_profiler で確実に取得
        out = run_command(["system_profiler", "SPDisplaysDataType"])
        current = {}
        for line in out.splitlines():
            line = line.strip()
            if "Chipset Model:" in line:
                if current:
                    gpus.append(current)
                current = {"GPU名": line.split(":", 1)[1].strip()}
            elif current:
                for keyword, label in [
                    ("VRAM",         "VRAM"),
                    ("Vendor:",      "ベンダー"),
                    ("Device ID:",   "デバイスID"),
                    ("Metal:",       "Metal対応"),
                    ("Resolution:",  "解像度"),
                    ("Displays:",    "接続ディスプレイ数"),
                ]:
                    if keyword in line:
                        current[label] = line.split(":", 1)[1].strip()
        if current:
            gpus.append(current)

    elif sys_name == "Linux":
        # ① lspci (pciutils)
        for lspci_cmd in [["lspci", "-vmm"], ["lspci", "-v"], ["lspci"]]:
            lspci_out = run_command(lspci_cmd)
            if lspci_out:
                break

        if lspci_out:
            current = {}
            in_gpu = False
            for line in lspci_out.splitlines():
                # -vmm 形式: "Class:	VGA..." などフラットに来る
                if re.match(r"^Class:\s+(VGA|3D|Display)", line):
                    if current:
                        gpus.append(current)
                    current = {}
                    in_gpu = True
                elif in_gpu and line.startswith("Device:"):
                    current["GPU名"] = line.split(":", 1)[1].strip()
                elif in_gpu and line.startswith("SVendor:"):
                    current["ベンダー"] = line.split(":", 1)[1].strip()
                elif in_gpu and line.startswith("Driver:"):
                    current["カーネルドライバ"] = line.split(":", 1)[1].strip()
                elif not line.strip() and in_gpu:
                    in_gpu = False

            # 通常 lspci 形式のフォールバック
            if not current and not gpus:
                for line in lspci_out.splitlines():
                    if any(k in line for k in ("VGA", "3D controller", "Display controller")):
                        parts = line.split(":", 2)
                        name = parts[2].strip() if len(parts) >= 3 else line
                        current = {"GPU名": name}
                    elif current and line.startswith("	"):
                        if "Kernel driver" in line:
                            current["カーネルドライバ"] = line.split(":", 1)[1].strip()
            if current:
                gpus.append(current)

        # ② /sys/class/drm フォールバック (lspci なし環境)
        if not gpus:
            drm_path = "/sys/class/drm"
            try:
                cards = [d for d in os.listdir(drm_path)
                         if re.match(r"^card\d+$", d)]
                for card in sorted(cards):
                    vendor_file = os.path.join(drm_path, card, "device", "vendor")
                    device_file = os.path.join(drm_path, card, "device", "device")
                    try:
                        vendor = open(vendor_file).read().strip()
                        device = open(device_file).read().strip()
                        gpus.append({"GPU名": f"{card} (vendor={vendor} device={device})"})
                    except Exception:
                        gpus.append({"GPU名": card})
            except Exception:
                pass

        # ③ nvidia-smi (NVIDIA GPU 専用)
        nsmi = run_command(["nvidia-smi",
                            "--query-gpu=name,memory.total,driver_version",
                            "--format=csv,noheader,nounits"])
        if nsmi:
            for line in nsmi.splitlines():
                parts = [p.strip() for p in line.split(",")]
                entry = {"GPU名": parts[0]} if parts else {}
                if len(parts) >= 2:
                    try:
                        entry["VRAM"] = f"{int(parts[1])/1024:.1f} GB"
                    except ValueError:
                        entry["VRAM"] = parts[1]
                if len(parts) >= 3:
                    entry["ドライババージョン"] = parts[2]
                if entry:
                    # 既存エントリと統合
                    matched = False
                    for g in gpus:
                        if parts[0].lower() in g.get("GPU名", "").lower():
                            g.update(entry)
                            matched = True
                            break
                    if not matched:
                        gpus.append(entry)

        if not gpus:
            gpus.append({"備考": "GPU情報を取得できませんでした (pciutils/lspci をインストールしてください)"})

    elif sys_name == "Windows":
        # PowerShell CIM で取得（WMIC 廃止対応）
        # 各プロパティを個別行で出力してパース
        ps_script = (
            "Get-CimInstance Win32_VideoController | ForEach-Object {"
            "  $ram = 0;"
            "  if ($_.AdapterRAM -and $_.AdapterRAM -gt 0) {"
            "    $ram = [math]::Round($_.AdapterRAM / 1073741824, 1)"
            "  };"
            "  $rx = if ($_.CurrentHorizontalResolution) { $_.CurrentHorizontalResolution } else { 0 };"
            "  $ry = if ($_.CurrentVerticalResolution)   { $_.CurrentVerticalResolution   } else { 0 };"
            "  $hz = if ($_.CurrentRefreshRate)          { $_.CurrentRefreshRate          } else { 0 };"
            "  Write-Output ('---GPU---');"
            "  Write-Output ('NAME:'     + $_.Name);"
            "  Write-Output ('VRAM:'     + $ram);"
            "  Write-Output ('DRIVER:'   + $_.DriverVersion);"
            "  Write-Output ('PROC:'     + $_.VideoProcessor);"
            "  Write-Output ('RESX:'     + $rx);"
            "  Write-Output ('RESY:'     + $ry);"
            "  Write-Output ('HZ:'       + $hz);"
            "  Write-Output ('STATUS:'   + $_.Status)"
            "}"
        )
        out = _ps_run(ps_script)
        current = {}
        for line in out.splitlines():
            line = line.strip()
            if line == "---GPU---":
                if current.get("GPU名"):
                    gpus.append(current)
                current = {}
            elif ":" in line:
                key, val = line.split(":", 1)
                val = val.strip()
                if key == "NAME" and val:
                    current["GPU名"] = val
                elif key == "VRAM":
                    try:
                        f = float(val)
                        current["VRAM"] = f"{f:.1f} GB" if f > 0 else "共有メモリ (統合GPU)"
                    except ValueError:
                        pass
                elif key == "DRIVER" and val:
                    current["ドライババージョン"] = val
                elif key == "PROC" and val:
                    current["ビデオプロセッサ"] = val
                elif key == "RESX" and val not in ("", "0"):
                    current.setdefault("_rx", val)
                elif key == "RESY" and val not in ("", "0"):
                    current.setdefault("_ry", val)
                elif key == "HZ" and val not in ("", "0"):
                    current.setdefault("_hz", val)
                elif key == "STATUS" and val:
                    current["ステータス"] = val
        if current.get("GPU名"):
            gpus.append(current)

        # 解像度を結合
        for g in gpus:
            rx = g.pop("_rx", None)
            ry = g.pop("_ry", None)
            hz = g.pop("_hz", None)
            if rx and ry:
                g["現在の解像度"] = f"{rx} x {ry}"
            if hz:
                g["リフレッシュレート"] = f"{hz} Hz"

        # フォールバック: Get-WmiObject (古い PowerShell 向け)
        if not gpus:
            fb_script = (
                "Get-WmiObject Win32_VideoController | ForEach-Object {"
                "  Write-Output ('---GPU---');"
                "  Write-Output ('NAME:'   + $_.Name);"
                "  Write-Output ('DRIVER:' + $_.DriverVersion)"
                "}"
            )
            out2 = _ps_run(fb_script)
            current = {}
            for line in out2.splitlines():
                line = line.strip()
                if line == "---GPU---":
                    if current.get("GPU名"):
                        gpus.append(current)
                    current = {}
                elif line.startswith("NAME:") and line[5:].strip():
                    current["GPU名"] = line[5:].strip()
                elif line.startswith("DRIVER:") and line[7:].strip():
                    current["ドライババージョン"] = line[7:].strip()
            if current.get("GPU名"):
                gpus.append(current)

        if not gpus:
            gpus.append({"備考": "GPU情報を取得できませんでした"})

    return gpus


# ── ウィジェット構築 ─────────────────────────────

class PCSpecApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PC スペックチェッカー")
        self.configure(bg=BG)
        self.geometry("1200x700")
        self.minsize(900, 500)
        self._spec_data = {}  # CSV出力用データ格納
        self._setup_fonts()
        self._build_ui()
        self._load_data_async()
        # OUIデータベースをバックグラウンドでロード開始（ステータスバーに進捗表示）
        _ensure_oui_db_async(
            on_status=lambda msg: self.after(0, lambda m=msg: self.oui_lbl.config(text=m))
        )

    def _setup_fonts(self):
        self.font_title  = ("Helvetica Neue", 22, "bold")
        self.font_head   = ("Helvetica Neue", 11, "bold")
        self.font_label  = ("Helvetica Neue", 10)
        self.font_val    = ("Menlo", 10) if platform.system() == "Darwin" else ("Consolas", 10)
        self.font_small  = ("Helvetica Neue", 9)

    def _build_ui(self):
        # ── ヘッダー
        header = tk.Frame(self, bg=PANEL, height=64)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        tk.Label(
            header, text="⬡  PC SPEC CHECKER",
            bg=PANEL, fg=ACCENT,
            font=self.font_title
        ).pack(side="left", padx=24, pady=14)

        self.scan_btn = tk.Button(
            header, text="🔍 LAN スキャン",
            bg="#7B3F00", fg=TEXT_MAIN,
            font=self.font_small,
            relief="flat", padx=12, pady=4,
            cursor="hand2",
            command=self._open_lan_scanner,
        )
        self.scan_btn.pack(side="right", padx=(0, 4), pady=14)

        self.xlsx_btn = tk.Button(
            header, text="📊 Excel出力",
            bg="#1d6f42", fg=TEXT_MAIN,
            font=self.font_small,
            relief="flat", padx=12, pady=4,
            cursor="hand2",
            command=self._export_xlsx,
            state="disabled"
        )
        self.xlsx_btn.pack(side="right", padx=(0, 4), pady=14)

        self.csv_btn = tk.Button(
            header, text="📥 CSV出力",
            bg=ACCENT2, fg=TEXT_MAIN,
            font=self.font_small,
            relief="flat", padx=12, pady=4,
            cursor="hand2",
            command=self._export_csv,
            state="disabled"
        )
        self.csv_btn.pack(side="right", padx=(0, 8), pady=14)

        self.status_lbl = tk.Label(
            header, text="読み込み中...",
            bg=PANEL, fg=TEXT_SUB,
            font=self.font_small
        )
        self.status_lbl.pack(side="right", padx=(4, 24))

        # OUI DB 専用ステータス（スペック収集と独立）
        self.oui_lbl = tk.Label(
            header, text="OUI DB: 待機中",
            bg=PANEL, fg="#7b4fff",
            font=self.font_small
        )
        self.oui_lbl.pack(side="right", padx=(4, 0))

        tk.Label(header, text="|", bg=PANEL, fg=BORDER,
                 font=self.font_small).pack(side="right", padx=2)

        # ── ノートブック
        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("TNotebook",
            background=BG, borderwidth=0, tabmargins=[0, 0, 0, 0])
        style.configure("TNotebook.Tab",
            background=PANEL, foreground=TEXT_SUB,
            padding=[18, 8], font=self.font_label, borderwidth=0)
        style.map("TNotebook.Tab",
            background=[("selected", BG)],
            foreground=[("selected", ACCENT)])
        style.configure("TFrame", background=BG)

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=0, pady=0)

        tab_names = ["機種情報", "OS", "CPU", "メモリ", "ストレージ", "GPU", "ネットワーク", "プリンター", "インストール済みアプリ"]
        self.tabs = {}
        for name in tab_names:
            frame = ttk.Frame(self.nb)
            self.nb.add(frame, text=f"  {name}  ")
            self.tabs[name] = frame

        # ── フッター
        footer = tk.Frame(self, bg=PANEL, height=28)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        if not HAS_PSUTIL:
            tk.Label(
                footer,
                text="ヒント: pip install psutil を実行するとより詳細な情報が表示されます",
                bg=PANEL, fg=YELLOW, font=self.font_small
            ).pack(side="left", padx=16, pady=4)

    # ── データ読み込み ─────────────────────────
    def _load_data_async(self):
        threading.Thread(target=self._collect_all, daemon=True).start()

    def _collect_all(self):
        self._update_status("機種情報を取得中...")
        machine_data = get_machine_info()
        self._spec_data["機種情報"] = machine_data
        self.after(0, lambda: self._render_kv(self.tabs["機種情報"], machine_data, "機種情報"))

        self._update_status("OS情報を取得中...")
        os_data = get_os_info()
        self._spec_data["OS"] = os_data
        self.after(0, lambda: self._render_kv(self.tabs["OS"], os_data))

        self._update_status("CPU情報を取得中...")
        cpu_data = get_cpu_info()
        self._spec_data["CPU"] = cpu_data
        self.after(0, lambda: self._render_cpu(self.tabs["CPU"], cpu_data))

        self._update_status("メモリ情報を取得中...")
        mem_data = get_memory_info()
        self._spec_data["メモリ"] = mem_data
        self.after(0, lambda: self._render_memory(self.tabs["メモリ"], mem_data))

        self._update_status("ストレージ情報を取得中...")
        disk_data = get_disk_info()
        self._spec_data["ストレージ"] = disk_data
        self.after(0, lambda: self._render_disks(self.tabs["ストレージ"], disk_data))

        self._update_status("GPU情報を取得中...")
        gpu_data = get_gpu_info()
        self._spec_data["GPU"] = gpu_data
        self.after(0, lambda: self._render_gpu(self.tabs["GPU"], gpu_data))

        self._update_status("ネットワーク情報を取得中...")
        net_data = get_network_info()
        self._spec_data["ネットワーク"] = net_data
        self.after(0, lambda: self._render_network(self.tabs["ネットワーク"], net_data))

        self._update_status("プリンター情報を取得中...")
        printer_data = get_printer_info()
        self._spec_data["プリンター"] = printer_data
        self.after(0, lambda: self._render_printers(self.tabs["プリンター"], printer_data))

        self._update_status("インストール済みアプリを取得中...")
        apps_data = get_installed_apps()
        self._spec_data["インストール済みアプリ"] = apps_data
        self.after(0, lambda: self._render_apps(self.tabs["インストール済みアプリ"], apps_data))

        self._update_status("完了")
        self.after(0, lambda: self.csv_btn.config(state="normal"))
        self.after(0, lambda: self.xlsx_btn.config(state="normal"))

    def _update_status(self, msg):
        self.after(0, lambda: self.status_lbl.config(text=msg))

    def _get_default_stem(self):
        """ファイル名のベース部分（拡張子なし）を生成"""
        hostname  = platform.node() or "unknown"
        serial    = self._spec_data.get("機種情報", {}).get("シリアルNo.", "") or ""
        safe_host = re.sub(r'[\\/:*?"<>|]', "_", hostname)
        safe_ser  = re.sub(r'[\\/:*?"<>|]', "_", serial)
        parts = ["pc_spec", safe_host]
        if safe_ser:
            parts.append(safe_ser)
        parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
        return "_".join(parts)

    def _open_lan_scanner(self):
        """LANスキャンウィンドウを開く"""
        import ipaddress, socket, concurrent.futures, time

        win = tk.Toplevel(self)
        win.title("LAN スキャン")
        win.configure(bg=BG)
        win.geometry("900x620")
        win.minsize(700, 450)

        # ── ツールバー ───────────────────────────────────────
        toolbar = tk.Frame(win, bg=PANEL, pady=6)
        toolbar.pack(fill="x", side="top")

        tk.Label(toolbar, text="対象ネットワーク:", bg=PANEL,
                 fg=TEXT_SUB, font=self.font_label).pack(side="left", padx=(12, 4))

        # 自 PC の IPv4 から CIDR を自動推定
        default_cidr = "192.168.1.0/24"
        try:
            if HAS_PSUTIL:
                import psutil as _ps
                for iface, addrs in _ps.net_if_addrs().items():
                    for a in addrs:
                        if a.family.name == "AF_INET" and not a.address.startswith("127."):
                            ip = ipaddress.IPv4Address(a.address)
                            mask = a.netmask or "255.255.255.0"
                            net = ipaddress.IPv4Network(f"{a.address}/{mask}", strict=False)
                            default_cidr = str(net)
                            raise StopIteration
        except StopIteration:
            pass
        except Exception:
            pass

        cidr_var = tk.StringVar(value=default_cidr)
        cidr_entry = tk.Entry(toolbar, textvariable=cidr_var,
                              bg="#1e2433", fg=TEXT_MAIN, insertbackground=TEXT_MAIN,
                              font=self.font_val, relief="flat", width=20)
        cidr_entry.pack(side="left", padx=4, ipady=4)

        timeout_var = tk.StringVar(value="300")
        tk.Label(toolbar, text="タイムアウト(ms):", bg=PANEL,
                 fg=TEXT_SUB, font=self.font_label).pack(side="left", padx=(12, 4))
        tk.Entry(toolbar, textvariable=timeout_var,
                 bg="#1e2433", fg=TEXT_MAIN, insertbackground=TEXT_MAIN,
                 font=self.font_val, relief="flat", width=6).pack(side="left", ipady=4)

        progress_lbl = tk.Label(toolbar, text="", bg=PANEL,
                                fg=TEXT_SUB, font=self.font_small)
        progress_lbl.pack(side="left", padx=16)

        # ── テーブル ─────────────────────────────────────────
        cols = ["IPアドレス", "ホスト名", "MACアドレス", "メーカー", "状態"]
        style = ttk.Style()
        style.configure("Scan.Treeview",
            background=BG, fieldbackground=BG,
            foreground=TEXT_VAL, rowheight=22,
            font=self.font_val, borderwidth=0)
        style.configure("Scan.Treeview.Heading",
            background=PANEL, foreground=ACCENT,
            font=self.font_head, relief="flat")
        style.map("Scan.Treeview",
            background=[("selected", "#1e2d4a")],
            foreground=[("selected", ACCENT)])

        table_frame = tk.Frame(win, bg=BG)
        table_frame.pack(fill="both", expand=True, padx=8, pady=(4, 0))

        tv = ttk.Treeview(table_frame, columns=cols, show="headings",
                          style="Scan.Treeview")
        col_widths = {"IPアドレス": 130, "ホスト名": 200, "MACアドレス": 150,
                      "メーカー": 180, "状態": 80}
        for col in cols:
            tv.heading(col, text=col,
                       command=lambda c=col: self._sort_tree(tv, c, False))
            tv.column(col, width=col_widths.get(col, 120), anchor="w", minwidth=60)

        vsb = ttk.Scrollbar(table_frame, orient="vertical",   command=tv.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tv.xview)
        tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        tv.tag_configure("up",      background="#0a1a10", foreground="#00e87a")
        tv.tag_configure("down",    background="#1a0a0a", foreground="#6b7a99")
        tv.tag_configure("self_pc", background="#0d1a2e", foreground="#00d4ff")

        # ── フッター ─────────────────────────────────────────
        footer = tk.Frame(win, bg=PANEL, pady=6)
        footer.pack(fill="x", side="bottom")

        count_lbl = tk.Label(footer, text="", bg=PANEL,
                             fg=TEXT_SUB, font=self.font_small)
        count_lbl.pack(side="left", padx=16)

        export_xlsx_btn = tk.Button(footer, text="📊 Excel出力", bg="#1d6f42", fg=TEXT_MAIN,
                                    font=self.font_small, relief="flat", padx=10, pady=3,
                                    cursor="hand2", state="disabled",
                                    command=lambda: self._export_scan_xlsx(tv, cols))
        export_xlsx_btn.pack(side="right", padx=4)

        export_btn = tk.Button(footer, text="📥 CSV出力", bg=ACCENT2, fg=TEXT_MAIN,
                               font=self.font_small, relief="flat", padx=10, pady=3,
                               cursor="hand2", state="disabled",
                               command=lambda: self._export_scan_csv(tv, cols))
        export_btn.pack(side="right", padx=4)

        scan_btn_inner = tk.Button(footer, text="▶ スキャン開始",
                                   bg=ACCENT, fg="#000000",
                                   font=self.font_small, relief="flat",
                                   padx=14, pady=3, cursor="hand2")
        scan_btn_inner.pack(side="right", padx=4)

        stop_flag = {"stop": False}

        def do_scan():
            stop_flag["stop"] = False
            scan_btn_inner.config(text="⏹ 停止", command=do_stop,
                                  bg=RED, fg=TEXT_MAIN)
            export_btn.config(state="disabled")
            tv.delete(*tv.get_children())
            count_lbl.config(text="スキャン中...")

            cidr_str = cidr_var.get().strip()
            try:
                timeout_ms = max(100, int(timeout_var.get()))
            except ValueError:
                timeout_ms = 300

            try:
                network = ipaddress.IPv4Network(cidr_str, strict=False)
            except ValueError:
                progress_lbl.config(text="CIDRが不正です")
                scan_btn_inner.config(text="▶ スキャン開始", command=do_scan,
                                      bg=ACCENT, fg="#000000")
                return

            hosts = list(network.hosts())
            total = len(hosts)
            found = [0]
            done  = [0]

            # ARP キャッシュを一度だけ取得（MACアドレス解決用）
            arp_table = _get_arp_table()

            # 自PCの IP → MAC マッピングを構築
            self_ips = set()
            self_ip_mac = {}   # {ip: (mac, vendor)}
            try:
                if HAS_PSUTIL:
                    import psutil as _ps2
                    import socket as _sock2
                    AF_INET = _sock2.AF_INET
                    MAC_FAM = {17}
                    if hasattr(_ps2, "AF_LINK"):
                        MAC_FAM.add(_ps2.AF_LINK)
                    for addrs in _ps2.net_if_addrs().values():
                        ipv4  = next((a.address for a in addrs if a.family == AF_INET
                                      and not a.address.startswith("127.")), None)
                        mac_r = next((a.address for a in addrs if a.family in MAC_FAM), None)
                        if ipv4:
                            self_ips.add(ipv4)
                        if ipv4 and mac_r:
                            mac_up = mac_r.replace("-", ":").upper()
                            self_ip_mac[ipv4] = (mac_up, _oui_vendor(mac_up))
                else:
                    import uuid as _uuid, socket as _sock3
                    raw = _uuid.getnode()
                    mac_fb = ":".join(
                        f"{(raw >> (8*i)) & 0xff:02X}" for i in reversed(range(6))
                    )
                    try:
                        my_ip = _sock3.gethostbyname(_sock3.gethostname())
                        self_ips.add(my_ip)
                        self_ip_mac[my_ip] = (mac_fb, _oui_vendor(mac_fb))
                    except Exception:
                        pass
            except Exception:
                pass

            def ping_host(ip_obj):
                if stop_flag["stop"]:
                    return None
                ip = str(ip_obj)
                is_self = ip in self_ips
                # 自PCはpingせずに確定で稼働中扱い
                alive = True if is_self else _ping(ip, timeout_ms)
                if not alive:
                    return None
                # ホスト名解決
                try:
                    hostname_r = socket.gethostbyaddr(ip)[0]
                except Exception:
                    hostname_r = ""
                # MAC: 自PCはpsutilから、他はARPキャッシュから
                if is_self and ip in self_ip_mac:
                    mac, vendor = self_ip_mac[ip]
                else:
                    mac, vendor = arp_table.get(ip, ("", ""))
                return (ip, hostname_r, mac, vendor, "自PC" if is_self else "稼働中", is_self)

            def update_ui(result):
                if result is None:
                    return
                ip, hostname_r, mac, vendor, status, is_self = result
                found[0] += 1
                tag = "self_pc" if is_self else "up"
                tv.insert("", "end",
                          values=(ip, hostname_r, mac, vendor, status),
                          tags=(tag,))
                # IPアドレス順にソート
                items = [(tv.set(k, "IPアドレス"), k) for k in tv.get_children("")]
                items.sort(key=lambda x: [int(p) for p in x[0].split(".") if p.isdigit()])
                for i, (_, k) in enumerate(items):
                    tv.move(k, "", i)

            def worker():
                max_workers = min(128, total)
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
                    futures = {ex.submit(ping_host, ip): ip for ip in hosts}
                    for fut in concurrent.futures.as_completed(futures):
                        if stop_flag["stop"]:
                            break
                        done[0] += 1
                        result = fut.result()
                        if result:
                            win.after(0, lambda r=result: update_ui(r))
                        pct = int(done[0] / total * 100)
                        win.after(0, lambda p=pct, d=done[0], t=total:
                                  progress_lbl.config(
                                      text=f"{d}/{t} ({p}%)"))

                win.after(0, scan_finished)

            def scan_finished():
                scan_btn_inner.config(text="▶ スキャン開始", command=do_scan,
                                      bg=ACCENT, fg="#000000")
                up = found[0]
                progress_lbl.config(text="完了")
                count_lbl.config(text=f"稼働中: {up} 台 / スキャン: {total} アドレス")
                export_btn.config(state="normal")
                export_xlsx_btn.config(state="normal")

            threading.Thread(target=worker, daemon=True).start()

        def do_stop():
            stop_flag["stop"] = True
            scan_btn_inner.config(text="▶ スキャン開始", command=do_scan,
                                  bg=ACCENT, fg="#000000")
            progress_lbl.config(text="停止しました")

        scan_btn_inner.config(command=do_scan)

    def _export_scan_csv(self, tv, cols):
        """スキャン結果をCSVに出力"""
        path = filedialog.asksaveasfilename(
            title="スキャン結果の保存先",
            defaultextension=".csv",
            initialfile=f"lan_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["出力日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                writer.writerow([])
                writer.writerow(cols)
                for iid in tv.get_children():
                    writer.writerow(tv.item(iid, "values"))
            messagebox.showinfo("完了", "スキャン結果を保存しました。\n\n" + path)
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    def _export_scan_xlsx(self, tv, cols):
        """LANスキャン結果を書式付きExcelに出力"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            if messagebox.askyesno(
                "openpyxl が見つかりません",
                "Excel出力には openpyxl が必要です。\n今すぐインストールしますか?\n\n  pip install openpyxl"
            ):
                import subprocess as _sp, sys as _sys
                try:
                    _sp.run([_sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
                    messagebox.showinfo("完了", "インストール完了。もう一度押してください。")
                except Exception as e:
                    messagebox.showerror("エラー", str(e))
            return

        default_name = f"lan_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Excelの保存先を選択",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excelファイル", "*.xlsx"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return

        try:
            # ── スタイル定数 ────────────────────────────────
            C = {
                "title_bg": "17375E", "title_fg": "FFFFFF",
                "info_bg":  "D6E4F0", "info_fg":  "17375E",
                "hdr_bg":   "2E75B6", "hdr_fg":   "FFFFFF",
                "odd_bg":   "FFFFFF", "even_bg":  "EBF3FB",
                "up_bg":    "E8F5E9", "up_fg":    "1B5E20",
                "self_bg":  "E3F2FD", "self_fg":  "0D47A1",
                "val_fg":   "17375E", "bdr":      "B8CCE4",
                "bdr_acc":  "2E75B6",
            }
            fn_name = "Yu Gothic UI" if platform.system() == "Windows" else "Helvetica Neue"

            def F(color, bold=False, size=10):
                return Font(name=fn_name, color=color, bold=bold, size=size)
            def Fill(c):
                return PatternFill("solid", fgColor=c)
            def Al(h="left", v="center", wrap=False):
                return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

            def full_border(ncols, ci, bg=None):
                bc = C["bdr"]
                l = Side(style="medium" if ci == 1      else "thin", color=bc)
                r = Side(style="medium" if ci == ncols  else "thin", color=bc)
                t = Side(style="thin",  color=bc)
                b = Side(style="thin",  color=bc)
                return Border(left=l, right=r, top=t, bottom=b)

            def hdr_border(ncols, ci):
                bc = C["bdr"]
                l = Side(style="medium" if ci == 1     else "thin",   color=bc)
                r = Side(style="medium" if ci == ncols else "thin",   color=bc)
                t = Side(style="medium", color=bc)
                b = Side(style="medium", color=C["bdr_acc"])
                return Border(left=l, right=r, top=t, bottom=b)

            # ── ワークブック作成 ─────────────────────────────
            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = "🌐 LANスキャン結果"
            ws.sheet_view.showGridLines = False
            ws.sheet_properties.tabColor = "1A472A"

            NC  = len(cols)
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            hostname = platform.node() or "unknown"

            # 列幅
            col_w = {"IPアドレス": 16, "ホスト名": 28, "MACアドレス": 20,
                     "メーカー": 24, "状態": 10}
            for ci, col in enumerate(cols, 1):
                ws.column_dimensions[get_column_letter(ci)].width = col_w.get(col, 18)

            row = 1
            # タイトル行
            ws.merge_cells(f"A{row}:{get_column_letter(NC)}{row}")
            tc = ws.cell(row=row, column=1, value="🌐  LAN スキャン結果")
            tc.fill = Fill(C["title_bg"]); tc.font = F(C["title_fg"], bold=True, size=14)
            tc.alignment = Al(h="center")
            for ci in range(1, NC + 1):
                cell = ws.cell(row=row, column=ci)
                cell.fill = Fill(C["title_bg"])
                l = Side(style="medium" if ci == 1  else "thin", color=C["title_bg"])
                r = Side(style="medium" if ci == NC else "thin", color=C["title_bg"])
                cell.border = Border(left=l, right=r,
                                     top=Side(style="medium", color=C["title_bg"]),
                                     bottom=Side(style="thin", color=C["bdr"]))
            ws.row_dimensions[row].height = 38
            row += 1

            # PC情報・スキャン日時行
            ws.merge_cells(f"A{row}:{get_column_letter(NC)}{row}")
            rows_data = list(tv.get_children())
            up_count  = sum(1 for iid in rows_data
                            if tv.item(iid, "values")[4] in ("稼働中", "自PC"))
            ic = ws.cell(row=row, column=1,
                         value=f"  PC名: {hostname}  │  スキャン日時: {now}"
                               f"  │  稼働ホスト: {up_count} 台 / {len(rows_data)} 件")
            ic.fill = Fill(C["info_bg"]); ic.font = F(C["info_fg"], size=9)
            ic.alignment = Al()
            for ci in range(1, NC + 1):
                cell = ws.cell(row=row, column=ci)
                cell.fill = Fill(C["info_bg"])
                l = Side(style="medium" if ci == 1  else "thin", color=C["bdr"])
                r = Side(style="medium" if ci == NC else "thin", color=C["bdr"])
                cell.border = Border(left=l, right=r,
                                     top=Side(style="thin",   color=C["bdr"]),
                                     bottom=Side(style="medium", color=C["bdr_acc"]))
            ws.row_dimensions[row].height = 20
            row += 1

            # 空白行
            ws.merge_cells(f"A{row}:{get_column_letter(NC)}{row}")
            for ci in range(1, NC + 1):
                ws.cell(row=row, column=ci).fill = Fill("FFFFFF")
            ws.row_dimensions[row].height = 6
            row += 1

            # 列ヘッダー行
            for ci, col in enumerate(cols, 1):
                hc = ws.cell(row=row, column=ci, value=col)
                hc.fill = Fill(C["hdr_bg"]); hc.font = F(C["hdr_fg"], bold=True, size=10)
                hc.alignment = Al(h="center"); hc.border = hdr_border(NC, ci)
            ws.row_dimensions[row].height = 22
            ws.freeze_panes = f"A{row + 1}"
            row += 1

            # データ行
            for idx, iid in enumerate(rows_data):
                values = tv.item(iid, "values")
                status = values[4] if len(values) > 4 else ""
                if status == "自PC":
                    bg = C["self_bg"]; fg = C["self_fg"]
                else:
                    bg = C["up_bg"]   if idx % 2 == 0 else C["even_bg"]
                    fg = C["val_fg"]

                for ci, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.fill = Fill(bg)
                    cell.font = F(fg, bold=(ci == 1), size=10)
                    cell.alignment = Al()
                    cell.border = full_border(NC, ci)
                ws.row_dimensions[row].height = 18
                row += 1

            # オートフィルター
            ws.auto_filter.ref = (
                f"A4:{get_column_letter(NC)}{row - 1}"
            )

            wb.save(path)
            messagebox.showinfo("Excel出力完了",
                                "スキャン結果を保存しました。\n\n" + path)
        except Exception as e:
            messagebox.showerror("エラー", "Excel出力に失敗しました。\n\n" + str(e))

    def _export_xlsx(self):
        """全スペック情報を見やすい書式付き Excel ファイルに出力する"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            if messagebox.askyesno(
                "openpyxl が見つかりません",
                "Excel出力には openpyxl が必要です。\n今すぐインストールしますか?\n\n  pip install openpyxl"
            ):
                import subprocess as _sp, sys as _sys
                try:
                    _sp.run([_sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
                    messagebox.showinfo("完了", "インストール完了。もう一度「Excel出力」を押してください。")
                except Exception as e:
                    messagebox.showerror("エラー", "インストール失敗:\n" + str(e))
            return

        stem = self._get_default_stem()
        path = filedialog.asksaveasfilename(
            title="Excelの保存先を選択",
            defaultextension=".xlsx",
            initialfile=stem + ".xlsx",
            filetypes=[("Excelファイル", "*.xlsx"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            hostname = platform.node() or "unknown"
            serial   = self._spec_data.get("機種情報", {}).get("シリアルNo.", "") or ""
            now_str  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # ═══════════════════════════════════════════
            # カラーパレット（ライト / ビジネス向け）
            # ═══════════════════════════════════════════
            C = {
                "hdr_bg"   : "17375E",   # 濃紺  ヘッダー背景
                "hdr_fg"   : "FFFFFF",   # 白    ヘッダー文字
                "acc_bg"   : "2E75B6",   # 中青  アクセント見出し背景
                "acc_fg"   : "FFFFFF",
                "sub_bg"   : "D6E4F0",   # 薄青  サブ見出し背景
                "sub_fg"   : "17375E",
                "odd_bg"   : "FFFFFF",   # 白    奇数行
                "even_bg"  : "EBF3FB",   # 極薄青 偶数行
                "key_fg"   : "404040",   # 濃グレー キー
                "val_fg"   : "17375E",   # 濃紺   値
                "title_bg" : "17375E",   # タイトル帯
                "title_fg" : "FFFFFF",
                "info_bg"  : "D6E4F0",   # PC情報帯
                "info_fg"  : "17375E",
                "bdr"      : "B8CCE4",   # 罫線
                "bdr_acc"  : "2E75B6",   # アクセント罫線
                "sum_cat"  : "1F497D",   # サマリーカテゴリ列
            }

            # ── スタイルヘルパー ───────────────────────
            def F(color=None, bold=False, size=10, italic=False):
                return Font(name="Yu Gothic UI" if platform.system()=="Windows" else "Helvetica Neue",
                            color=color or C["key_fg"], bold=bold, size=size, italic=italic)

            def Fill(hex_c):
                return PatternFill("solid", fgColor=hex_c)

            def Al(h="left", v="center", wrap=False, indent=0):
                return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

            def Bdr(color=None, style="thin"):
                """全辺均一罫線"""
                s = Side(style=style, color=color or C["bdr"])
                return Border(left=s, right=s, top=s, bottom=s)

            def BdrOuter(ncols, col):
                """左端・右端・中間で外枠/内線を使い分け"""
                bdr_c = C["bdr"]
                left  = Side(style="medium" if col == 1     else "thin", color=bdr_c)
                right = Side(style="medium" if col == ncols else "thin", color=bdr_c)
                top   = Side(style="thin",   color=bdr_c)
                bot   = Side(style="thin",   color=bdr_c)
                return Border(left=left, right=right, top=top, bottom=bot)

            def BdrAccBottom(ncols=1, col=1):
                """下辺をアクセント色mediumにした罫線"""
                bdr_c = C["bdr"]
                left  = Side(style="medium" if col == 1     else "thin", color=bdr_c)
                right = Side(style="medium" if col == ncols else "thin", color=bdr_c)
                top   = Side(style="thin",   color=bdr_c)
                bot   = Side(style="medium", color=C["bdr_acc"])
                return Border(left=left, right=right, top=top, bottom=bot)

            def apply_merge_border(ws, row, ncols, fill_obj, bdr_func, *args):
                """マージセルの全列にきちんと罫線を適用するユーティリティ"""
                for ci in range(1, ncols + 1):
                    c = ws.cell(row=row, column=ci)
                    c.fill   = fill_obj
                    c.border = bdr_func(ncols, ci, *args)

            def set_col_widths(ws, widths):
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

            def freeze(ws, cell="A5"):
                ws.freeze_panes = cell

            def make_sheet(title, tab_color="17375E"):
                ws = wb.create_sheet(title=title)
                ws.sheet_view.showGridLines = False
                ws.sheet_properties.tabColor = tab_color
                return ws

            # ── 共通ブロック描画 ──────────────────────

            def draw_title_bar(ws, row, text, ncols=3):
                ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
                c = ws.cell(row=row, column=1, value=text)
                c.fill = Fill(C["title_bg"]); c.font = F(C["title_fg"], bold=True, size=14)
                c.alignment = Al(h="center")
                # マージ全列に外枠罫線
                for ci in range(1, ncols + 1):
                    cell = ws.cell(row=row, column=ci)
                    cell.fill = Fill(C["title_bg"])
                    left  = Side(style="medium" if ci == 1     else "thin", color=C["title_bg"])
                    right = Side(style="medium" if ci == ncols else "thin", color=C["title_bg"])
                    top   = Side(style="medium", color=C["title_bg"])
                    bot   = Side(style="thin",   color=C["bdr"])
                    cell.border = Border(left=left, right=right, top=top, bottom=bot)
                ws.row_dimensions[row].height = 38
                return row + 1

            def draw_info_bar(ws, row, ncols=3):
                ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
                serial_part = f"  │  シリアルNo: {serial}" if serial else ""
                c = ws.cell(row=row, column=1,
                            value=f"  PC名: {hostname}{serial_part}  │  出力日時: {now_str}")
                c.fill = Fill(C["info_bg"]); c.font = F(C["info_fg"], bold=False, size=9)
                c.alignment = Al()
                for ci in range(1, ncols + 1):
                    cell = ws.cell(row=row, column=ci)
                    cell.fill = Fill(C["info_bg"])
                    left  = Side(style="medium" if ci == 1     else "thin", color=C["bdr"])
                    right = Side(style="medium" if ci == ncols else "thin", color=C["bdr"])
                    top   = Side(style="thin",   color=C["bdr"])
                    bot   = Side(style="medium", color=C["bdr_acc"])
                    cell.border = Border(left=left, right=right, top=top, bottom=bot)
                ws.row_dimensions[row].height = 20
                return row + 1

            def draw_col_header(ws, row, labels, ncols=None):
                """列ヘッダー: 各セルに独立した罫線（外枠medium, 内線thin）"""
                n = ncols or len(labels)
                for ci, lbl in enumerate(labels, 1):
                    c = ws.cell(row=row, column=ci, value=lbl)
                    c.fill = Fill(C["hdr_bg"]); c.font = F(C["hdr_fg"], bold=True, size=10)
                    c.alignment = Al(h="center")
                    left  = Side(style="medium" if ci == 1  else "thin", color=C["hdr_bg"])
                    right = Side(style="medium" if ci == n  else "thin", color=C["hdr_bg"])
                    top   = Side(style="medium", color=C["hdr_bg"])
                    bot   = Side(style="medium", color=C["bdr_acc"])
                    c.border = Border(left=left, right=right, top=top, bottom=bot)
                ws.row_dimensions[row].height = 22
                return row + 1

            def draw_section_header(ws, row, text, ncols=3):
                ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
                c = ws.cell(row=row, column=1, value=f"  ◆  {text}")
                c.fill = Fill(C["acc_bg"]); c.font = F(C["acc_fg"], bold=True, size=11)
                c.alignment = Al()
                for ci in range(1, ncols + 1):
                    cell = ws.cell(row=row, column=ci)
                    cell.fill = Fill(C["acc_bg"])
                    left  = Side(style="medium" if ci == 1     else "thin", color=C["bdr"])
                    right = Side(style="medium" if ci == ncols else "thin", color=C["bdr"])
                    top   = Side(style="thin",   color=C["bdr"])
                    bot   = Side(style="medium", color=C["bdr_acc"])
                    cell.border = Border(left=left, right=right, top=top, bottom=bot)
                ws.row_dimensions[row].height = 24
                return row + 1

            def draw_sub_header(ws, row, text, ncols=3):
                ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
                c = ws.cell(row=row, column=1, value=f"    ▸  {text}")
                c.fill = Fill(C["sub_bg"]); c.font = F(C["sub_fg"], bold=True, size=10)
                c.alignment = Al()
                for ci in range(1, ncols + 1):
                    cell = ws.cell(row=row, column=ci)
                    cell.fill = Fill(C["sub_bg"])
                    left  = Side(style="medium" if ci == 1     else "thin", color=C["bdr"])
                    right = Side(style="medium" if ci == ncols else "thin", color=C["bdr"])
                    top   = Side(style="thin",   color=C["bdr"])
                    bot   = Side(style="thin",   color=C["bdr_acc"])
                    cell.border = Border(left=left, right=right, top=top, bottom=bot)
                ws.row_dimensions[row].height = 20
                return row + 1

            def draw_kv(ws, row, key, value, idx=0, ncols=3):
                """キー列(A) + 値列(B〜最終列マージ) のデータ行"""
                bg = C["odd_bg"] if idx % 2 == 0 else C["even_bg"]
                # キー列 (A)
                kc = ws.cell(row=row, column=1, value=key)
                kc.fill = Fill(bg); kc.font = F(C["key_fg"], size=10)
                kc.alignment = Al(indent=1)
                kc.border = Border(
                    left   = Side(style="medium", color=C["bdr"]),
                    right  = Side(style="thin",   color=C["bdr"]),
                    top    = Side(style="thin",   color=C["bdr"]),
                    bottom = Side(style="thin",   color=C["bdr"]),
                )
                # 値列 (B〜ncols): マージしつつ全セルに罫線
                if ncols >= 2:
                    ws.merge_cells(f"B{row}:{get_column_letter(ncols)}{row}")
                    for ci in range(2, ncols + 1):
                        cell = ws.cell(row=row, column=ci)
                        cell.fill = Fill(bg)
                        left  = Side(style="thin",   color=C["bdr"])
                        right = Side(style="medium" if ci == ncols else "thin", color=C["bdr"])
                        top   = Side(style="thin",   color=C["bdr"])
                        bot   = Side(style="thin",   color=C["bdr"])
                        cell.border = Border(left=left, right=right, top=top, bottom=bot)
                    vc = ws.cell(row=row, column=2,
                                 value=str(value) if value else "—")
                    vc.font = F(C["val_fg"], bold=True, size=10)
                    vc.alignment = Al(wrap=True)
                ws.row_dimensions[row].height = 18
                return row + 1

            def draw_empty(ws, row, ncols=3):
                """区切り空白行: 罫線なし・背景白"""
                if ncols > 1:
                    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
                for ci in range(1, ncols + 1):
                    cell = ws.cell(row=row, column=ci)
                    cell.fill   = Fill("FFFFFF")
                    cell.border = Border()   # 罫線なし（前行の下罫線だけ残す）
                ws.row_dimensions[row].height = 6
                return row + 1

            cat_icons = {
                "機種情報": "🖥 機種情報", "OS": "💻 OS",
                "CPU": "⚡ CPU",      "メモリ": "🧠 メモリ",
                "ストレージ": "💾 ストレージ", "GPU": "🎮 GPU",
                "ネットワーク": "🌐 ネットワーク",
                "インストール済みアプリ": "📦 インストール済みアプリ",
                "プリンター": "🖨 プリンター",
            }
            tab_colors = {
                "機種情報": "17375E", "OS": "375623", "CPU": "7B3F00",
                "メモリ": "4A235A", "ストレージ": "78290F",
                "GPU": "1C3A5E", "ネットワーク": "1A472A",
                "インストール済みアプリ": "5C3317",
                "プリンター": "3D1A78",
            }

            # ═══════════════════════════════════════════
            # ① サマリーシート
            # ═══════════════════════════════════════════
            ws_s = make_sheet("📋 サマリー", "17375E")
            SNC = 3  # サマリーは3列固定
            set_col_widths(ws_s, [20, 28, 42])

            row = draw_title_bar(ws_s, 1, "🖥  PC SPEC CHECKER  ─  スペック一覧", ncols=SNC)
            row = draw_info_bar(ws_s, row,                                          ncols=SNC)
            row = draw_empty(ws_s, row,                                             ncols=SNC)
            row = draw_col_header(ws_s, row, ["カテゴリ", "項目", "値"])
            freeze(ws_s, f"A{row}")

            idx = 0
            prev_cat = None
            for cat, data in self._spec_data.items():
                rows_to_write = []
                if isinstance(data, dict):
                    for k, v in data.items():
                        rows_to_write.append((cat, k, str(v) if v else "—"))
                elif isinstance(data, list):
                    for item in data:
                        if not isinstance(item, dict):
                            continue
                        sub = (item.get("マウント") or item.get("GPU名")
                               or item.get("インターフェース") or item.get("備考") or "")
                        lbl = f"{cat}  ({sub})" if sub else cat
                        for k, v in item.items():
                            rows_to_write.append((lbl, k, str(v) if v else "—"))

                for cat_lbl, k, v in rows_to_write:
                    bg = C["odd_bg"] if idx % 2 == 0 else C["even_bg"]
                    # カテゴリ列
                    cc = ws_s.cell(row=row, column=1, value=cat_lbl if cat_lbl != prev_cat else "")
                    cc.fill = Fill(bg); cc.font = F(C["sum_cat"], bold=(cat_lbl != prev_cat), size=9)
                    cc.alignment = Al(indent=1); cc.border = Bdr()
                    prev_cat = cat_lbl
                    # キー列
                    kc = ws_s.cell(row=row, column=2, value=k)
                    kc.fill = Fill(bg); kc.font = F(C["key_fg"], size=10)
                    kc.alignment = Al(indent=1); kc.border = Bdr()
                    # 値列
                    vc = ws_s.cell(row=row, column=3, value=v)
                    vc.fill = Fill(bg); vc.font = F(C["val_fg"], bold=True, size=10)
                    vc.alignment = Al(wrap=True); vc.border = Bdr()
                    ws_s.row_dimensions[row].height = 18
                    row += 1; idx += 1

            # ═══════════════════════════════════════════
            # ② カテゴリ別シート
            # ═══════════════════════════════════════════
            for cat, data in self._spec_data.items():
                icon   = cat_icons.get(cat, cat)
                tcolor = tab_colors.get(cat, "17375E")
                ws     = make_sheet(icon, tcolor)

                # ── dict 系（OS/CPU/メモリ/機種情報）: 2列構成（項目・値）
                if isinstance(data, dict):
                    NC = 2
                    set_col_widths(ws, [30, 50])
                    r = draw_title_bar(ws, 1, icon, ncols=NC)
                    r = draw_info_bar(ws, r,         ncols=NC)
                    r = draw_empty(ws, r,             ncols=NC)
                    r = draw_section_header(ws, r, cat, ncols=NC)
                    r = draw_col_header(ws, r, ["項目", "値"])
                    for i, (k, v) in enumerate(data.items()):
                        r = draw_kv(ws, r, k, v, i, ncols=NC)
                    freeze(ws, f"A{r - len(data)}")

                elif isinstance(data, list):
                    if cat == "インストール済みアプリ":
                        # アプリは多列テーブル形式
                        priority = ["アプリ名", "パッケージ名", "バージョン",
                                    "発行元", "インストール日", "管理", "場所"]
                        present = set()
                        for item in data:
                            if isinstance(item, dict):
                                present |= set(item.keys())
                        present.discard("備考")
                        app_cols = ([c for c in priority if c in present] +
                                    [c for c in present if c not in priority])
                        NC = len(app_cols)
                        col_w = {"アプリ名": 40, "パッケージ名": 40,
                                 "バージョン": 16, "発行元": 28,
                                 "インストール日": 14, "管理": 10, "場所": 20}
                        set_col_widths(ws, [col_w.get(c, 18) for c in app_cols])
                        r = draw_title_bar(ws, 1, icon, ncols=NC)
                        r = draw_info_bar(ws, r,         ncols=NC)
                        r = draw_empty(ws, r,             ncols=NC)
                        # 列ヘッダー
                        for ci, h in enumerate(app_cols, 1):
                            hc = ws.cell(row=r, column=ci, value=h)
                            hc.fill = Fill(C["hdr_bg"])
                            hc.font = F(C["hdr_fg"], bold=True, size=10)
                            hc.alignment = Al(h="center")
                            hc.border = Bdr(C["hdr_bg"])
                        ws.row_dimensions[r].height = 22
                        freeze(ws, f"A{r + 1}")
                        r += 1
                        # データ行
                        for ai, item in enumerate(data):
                            if not isinstance(item, dict):
                                continue
                            bg = C["odd_bg"] if ai % 2 == 0 else C["even_bg"]
                            for ci, col in enumerate(app_cols, 1):
                                val = str(item.get(col, "")) if item.get(col) else "—"
                                c = ws.cell(row=r, column=ci, value=val)
                                c.fill = Fill(bg)
                                c.font = F(C["val_fg"] if ci > 1 else C["sum_cat"],
                                           bold=(ci == 1), size=9)
                                c.alignment = Al(indent=1)
                                c.border = Bdr()
                            ws.row_dimensions[r].height = 16
                            r += 1

                    else:
                        # ストレージ / GPU / ネットワーク / プリンター: 2列構成
                        NC = 2
                        set_col_widths(ws, [30, 50])
                        r = draw_title_bar(ws, 1, icon, ncols=NC)
                        r = draw_info_bar(ws, r,         ncols=NC)
                        r = draw_empty(ws, r,             ncols=NC)
                        r = draw_col_header(ws, r, ["項目", "値"])
                        freeze(ws, f"A{r}")
                        for item in data:
                            if not isinstance(item, dict):
                                continue
                            sub = (item.get("マウント") or item.get("GPU名")
                                   or item.get("プリンター名")
                                   or item.get("インターフェース") or item.get("備考") or "")
                            r = draw_sub_header(ws, r, sub or cat, ncols=NC)
                            for i, (k, v) in enumerate(item.items()):
                                r = draw_kv(ws, r, k, v, i, ncols=NC)
                            r = draw_empty(ws, r, ncols=NC)

            wb.save(path)
            messagebox.showinfo("Excel出力完了", "スペック情報を保存しました。\n\n" + path)

        except Exception as e:
            messagebox.showerror("エラー", "Excel出力に失敗しました。\n\n" + str(e))

    def _export_csv(self):
        """全スペック情報を CSV ファイルに出力する"""
        hostname = platform.node() or "unknown"
        default_name = self._get_default_stem() + ".csv"
        path = filedialog.asksaveasfilename(
            title="CSVの保存先を選択",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return  # キャンセル

        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["PC名", hostname])
                writer.writerow(["出力日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                writer.writerow([])
                writer.writerow(["カテゴリ", "項目", "値"])
                writer.writerow([])

                for category, data in self._spec_data.items():
                    if isinstance(data, dict):
                        # OS, CPU, メモリ, 機種情報
                        for key, val in data.items():
                            writer.writerow([category, key, val])
                        writer.writerow([])

                    elif isinstance(data, list):
                        for i, item in enumerate(data):
                            if not isinstance(item, dict):
                                continue
                            # ストレージ: マウント名、GPU: GPU名、ネットワーク: インターフェース名を小見出しに
                            subtitle = (
                                item.get("マウント")
                                or item.get("GPU名")
                                or item.get("インターフェース")
                                or item.get("備考")
                                or f"#{i+1}"
                            )
                            for key, val in item.items():
                                writer.writerow([f"{category} ({subtitle})", key, val])
                            writer.writerow([])

            messagebox.showinfo(
                "CSV出力完了",
                "スペック情報を保存しました。\n\n" + path
            )
        except Exception as e:
            messagebox.showerror("エラー", "CSV出力に失敗しました。\n\n" + str(e))

    # ── レンダリング ────────────────────────────
    def _scrollable(self, parent):
        canvas = tk.Canvas(parent, bg=BG, highlightthickness=0, bd=0)
        sb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        style = ttk.Style()
        style.configure("Vertical.TScrollbar",
            background=BORDER, troughcolor=PANEL,
            arrowcolor=TEXT_SUB, relief="flat", borderwidth=0)
        frame = tk.Frame(canvas, bg=BG)
        frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        win_id = canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.bind("<Configure>",
            lambda e: canvas.itemconfig(win_id, width=e.width))
        canvas.pack(side="left",  fill="both", expand=True)
        sb.pack   (side="right", fill="y")
        canvas.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))
        return frame

    def _section_label(self, parent, text):
        f = tk.Frame(parent, bg=BG)
        f.pack(fill="x", padx=24, pady=(20, 6))
        tk.Label(f, text=text.upper(), bg=BG,
                 fg=ACCENT, font=self.font_head).pack(side="left")
        tk.Frame(f, bg=BORDER, height=1).pack(
            side="left", fill="x", expand=True, padx=(10, 0))

    def _kv_row(self, parent, label, value, alt=False):
        bg = "#111520" if alt else BG
        row = tk.Frame(parent, bg=bg)
        row.pack(fill="x", padx=24, pady=1)

        lbl = tk.Label(row, text=label, bg=bg,
                       fg=TEXT_SUB, font=self.font_label, width=22, anchor="w")
        lbl.pack(side="left", padx=(0, 8), pady=5)

        val = tk.Label(row, text=value or "—", bg=bg,
                       fg=TEXT_VAL, font=self.font_val, anchor="w")
        val.pack(side="left", pady=5)

    def _render_kv(self, tab, data: dict, section="情報"):
        frame = self._scrollable(tab)
        if not data:
            self._section_label(frame, section)
            tk.Label(frame, text="情報を取得できませんでした",
                     bg=BG, fg=TEXT_SUB, font=self.font_label).pack(padx=32, pady=10)
            return
        self._section_label(frame, section if section != "情報" else list(data.keys())[0] if len(data) == 1 else "情報")
        for i, (k, v) in enumerate(data.items()):
            self._kv_row(frame, k, v, alt=(i % 2 == 1))

    def _render_cpu(self, tab, data: dict):
        frame = self._scrollable(tab)
        self._section_label(frame, "プロセッサー")
        for i, (k, v) in enumerate(data.items()):
            self._kv_row(frame, k, v, alt=(i % 2 == 1))

        if HAS_PSUTIL:
            self._section_label(frame, "リアルタイム使用率")
            bar_frame = tk.Frame(frame, bg=BG)
            bar_frame.pack(fill="x", padx=24, pady=8)
            self._draw_cpu_bars(bar_frame)

    def _draw_cpu_bars(self, parent):
        try:
            percents = psutil.cpu_percent(percpu=True, interval=0.5)
        except Exception:
            return
        cols = min(4, len(percents))
        for i, pct in enumerate(percents):
            col = i % cols
            row = i // cols
            cell = tk.Frame(parent, bg=BG)
            cell.grid(row=row, column=col, padx=8, pady=4, sticky="ew")
            parent.columnconfigure(col, weight=1)

            lbl = tk.Label(cell, text=f"Core {i}",
                           bg=BG, fg=TEXT_SUB, font=self.font_small)
            lbl.pack(anchor="w")

            bar_bg = tk.Frame(cell, bg=BORDER, height=6)
            bar_bg.pack(fill="x")
            bar_bg.update_idletasks()
            w = bar_bg.winfo_width() or 120
            fill_w = max(4, int(w * pct / 100))
            color = GREEN if pct < 60 else YELLOW if pct < 85 else RED
            bar_fill = tk.Frame(bar_bg, bg=color, height=6, width=fill_w)
            bar_fill.place(x=0, y=0)

            tk.Label(cell, text=f"{pct:.0f}%",
                     bg=BG, fg=TEXT_VAL, font=self.font_small).pack(anchor="e")

    def _render_memory(self, tab, data: dict):
        frame = self._scrollable(tab)
        self._section_label(frame, "物理メモリ")
        for i, (k, v) in enumerate(data.items()):
            self._kv_row(frame, k, v, alt=(i % 2 == 1))

        if HAS_PSUTIL:
            mem = psutil.virtual_memory()
            self._section_label(frame, "使用状況")
            self._draw_usage_bar(frame, "RAM", mem.percent, GREEN if mem.percent < 70 else YELLOW if mem.percent < 90 else RED)

            swap = psutil.swap_memory()
            if swap.total > 0:
                self._draw_usage_bar(frame, "SWAP", swap.percent, ACCENT2)

    def _draw_usage_bar(self, parent, label, pct, color):
        f = tk.Frame(parent, bg=BG)
        f.pack(fill="x", padx=24, pady=6)
        tk.Label(f, text=label, bg=BG,
                 fg=TEXT_SUB, font=self.font_label, width=10, anchor="w").pack(side="left")
        bar_bg = tk.Frame(f, bg=BORDER, height=16)
        bar_bg.pack(side="left", fill="x", expand=True, padx=(8, 8))
        bar_bg.update_idletasks()
        bar_bg.pack_propagate(False)
        # Use canvas for reliable pixel width
        bar_canvas = tk.Canvas(bar_bg, bg=BORDER, height=16, highlightthickness=0, bd=0)
        bar_canvas.pack(fill="both", expand=True)
        bar_canvas.update_idletasks()
        w = bar_canvas.winfo_width() or 300
        fill_w = max(4, int(w * pct / 100))
        bar_canvas.create_rectangle(0, 0, fill_w, 16, fill=color, outline="")
        tk.Label(f, text=f"{pct:.1f}%", bg=BG,
                 fg=TEXT_VAL, font=self.font_val, width=6).pack(side="right")

    def _render_disks(self, tab, disks: list):
        frame = self._scrollable(tab)
        if not disks:
            self._section_label(frame, "ストレージ")
            tk.Label(frame, text="情報を取得できませんでした",
                     bg=BG, fg=TEXT_SUB, font=self.font_label).pack(padx=32, pady=10)
            return

        for disk in disks:
            mount = disk.get("マウント") or disk.get("デバイス", "ドライブ")
            self._section_label(frame, f"ドライブ: {mount}")
            for i, (k, v) in enumerate(disk.items()):
                self._kv_row(frame, k, v, alt=(i % 2 == 1))

            pct_str = disk.get("使用率", "0 %").replace("%", "").strip()
            try:
                pct = float(pct_str)
                color = GREEN if pct < 70 else YELLOW if pct < 90 else RED
                self._draw_usage_bar(frame, "使用率", pct, color)
            except ValueError:
                pass


    def _render_gpu(self, tab, gpus: list):
        frame = self._scrollable(tab)
        if not gpus:
            self._section_label(frame, "GPU")
            tk.Label(frame, text="情報を取得できませんでした",
                     bg=BG, fg=TEXT_SUB, font=self.font_label).pack(padx=32, pady=10)
            return
        for gpu in gpus:
            name = gpu.get("GPU名") or gpu.get("備考", "GPU")
            self._section_label(frame, name)
            items = [(k, v) for k, v in gpu.items() if k != "GPU名"]
            for i, (k, v) in enumerate(items):
                self._kv_row(frame, k, v, alt=(i % 2 == 1))

    def _render_network(self, tab, interfaces: list):
        frame = self._scrollable(tab)
        if not interfaces:
            self._section_label(frame, "ネットワーク")
            tk.Label(frame, text="情報を取得できませんでした",
                     bg=BG, fg=TEXT_SUB, font=self.font_label).pack(padx=32, pady=10)
            return

        for iface_data in interfaces:
            name = iface_data.get("インターフェース", "不明")
            self._section_label(frame, f"インターフェース: {name}")
            items = [(k, v) for k, v in iface_data.items() if k != "インターフェース"]
            for i, (k, v) in enumerate(items):
                # MACアドレスは強調表示
                if k == "MACアドレス":
                    row_bg = "#0d1a22"
                    lbl_fg = ACCENT
                    val_fg = ACCENT
                    row = tk.Frame(frame, bg=row_bg)
                    row.pack(fill="x", padx=24, pady=1)
                    tk.Label(row, text=k, bg=row_bg, fg=lbl_fg,
                             font=self.font_label, width=22, anchor="w").pack(
                        side="left", padx=(0, 8), pady=5)
                    tk.Label(row, text=v or "—", bg=row_bg, fg=val_fg,
                             font=self.font_val, anchor="w").pack(side="left", pady=5)
                else:
                    self._kv_row(frame, k, v, alt=(i % 2 == 1))


    def _render_printers(self, tab, printers: list):
        """プリンター一覧を表示"""
        frame = self._scrollable(tab)

        if not printers or (len(printers) == 1 and "備考" in printers[0]):
            self._section_label(frame, "プリンター")
            tk.Label(frame, text="プリンターが見つかりませんでした",
                     bg=BG, fg=TEXT_SUB, font=self.font_label).pack(padx=32, pady=16)
            return

        self._section_label(frame, f"プリンター ({len(printers)} 台)")

        for i, p in enumerate(printers):
            name = p.get("プリンター名", f"プリンター #{i+1}")
            is_default = p.get("既定", "") == "はい"

            # プリンターカード
            card_bg = "#0d1a2e" if i % 2 == 0 else "#111b30"
            card = tk.Frame(frame, bg=card_bg, pady=2)
            card.pack(fill="x", padx=16, pady=4)

            # タイトル行
            title_f = tk.Frame(card, bg=card_bg)
            title_f.pack(fill="x", padx=12, pady=(8, 4))

            icon = "🖨"
            title_text = f"{icon}  {name}"
            if is_default:
                title_text += "  ★ 既定"
            tk.Label(title_f, text=title_text,
                     bg=card_bg, fg=ACCENT if is_default else TEXT_MAIN,
                     font=self.font_head).pack(side="left")

            # 詳細情報
            detail_f = tk.Frame(card, bg=card_bg)
            detail_f.pack(fill="x", padx=24, pady=(0, 8))
            skip = {"プリンター名", "既定"}
            for j, (k, v) in enumerate(p.items()):
                if k in skip or not v:
                    continue
                row = tk.Frame(detail_f, bg=card_bg)
                row.pack(fill="x", pady=1)
                # 状態に色付け
                if k == "状態":
                    col = GREEN if "待機" in v else YELLOW if "印刷中" in v else RED
                else:
                    col = TEXT_VAL
                tk.Label(row, text=k, bg=card_bg, fg=TEXT_SUB,
                         font=self.font_label, width=18, anchor="w").pack(side="left")
                tk.Label(row, text=v, bg=card_bg, fg=col,
                         font=self.font_val, anchor="w").pack(side="left")

    def _render_apps(self, tab, apps: list):
        """インストール済みアプリをリスト＋検索ボックスで表示"""
        # 外枠
        outer = tk.Frame(tab, bg=BG)
        outer.pack(fill="both", expand=True)

        # ── 検索バー ──
        search_bar = tk.Frame(outer, bg=PANEL, pady=6)
        search_bar.pack(fill="x", side="top")

        tk.Label(search_bar, text="🔍  検索:", bg=PANEL,
                 fg=TEXT_SUB, font=self.font_label).pack(side="left", padx=(16, 4))
        search_var = tk.StringVar()
        entry = tk.Entry(search_bar, textvariable=search_var,
                         bg="#1e2433", fg=TEXT_MAIN, insertbackground=TEXT_MAIN,
                         font=self.font_val, relief="flat", width=32)
        entry.pack(side="left", padx=4, ipady=4)

        count_lbl = tk.Label(search_bar, text="", bg=PANEL,
                             fg=TEXT_SUB, font=self.font_small)
        count_lbl.pack(side="left", padx=12)

        # ── テーブル (Treeview) ──
        cols = self._app_columns(apps)
        tree_frame = tk.Frame(outer, bg=BG)
        tree_frame.pack(fill="both", expand=True)

        style = ttk.Style()
        style.configure("Apps.Treeview",
            background=BG, fieldbackground=BG,
            foreground=TEXT_VAL, rowheight=22,
            font=self.font_val, borderwidth=0)
        style.configure("Apps.Treeview.Heading",
            background=PANEL, foreground=ACCENT,
            font=self.font_head, relief="flat")
        style.map("Apps.Treeview",
            background=[("selected", "#1e2d4a")],
            foreground=[("selected", ACCENT)])

        tv = ttk.Treeview(tree_frame, columns=cols, show="headings",
                          style="Apps.Treeview")

        # カラム幅設定
        col_widths = {"アプリ名": 280, "パッケージ名": 280,
                      "バージョン": 120, "発行元": 180,
                      "インストール日": 110, "管理": 80, "場所": 160}
        for col in cols:
            w = col_widths.get(col, 140)
            tv.heading(col, text=col,
                       command=lambda c=col: self._sort_tree(tv, c, False))
            tv.column(col, width=w, anchor="w", minwidth=60)

        # スクロールバー
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tv.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tv.xview)
        tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        # 行データを挿入
        all_rows = []
        for app in apps:
            row_vals = tuple(str(app.get(c, "")) for c in cols)
            all_rows.append(row_vals)

        def populate(rows):
            tv.delete(*tv.get_children())
            for i, vals in enumerate(rows):
                tag = "odd" if i % 2 == 0 else "even"
                tv.insert("", "end", values=vals, tags=(tag,))
            tv.tag_configure("odd",  background="#0d1525")
            tv.tag_configure("even", background="#111b30")
            count_lbl.config(
                text=f"{len(rows)} / {len(all_rows)} 件")

        populate(all_rows)

        # 検索フィルター
        def on_search(*_):
            q = search_var.get().lower()
            if not q:
                populate(all_rows)
            else:
                filtered = [r for r in all_rows
                            if any(q in str(v).lower() for v in r)]
                populate(filtered)

        search_var.trace_add("write", on_search)

    def _app_columns(self, apps):
        """実際に値がある列だけを返す"""
        priority = ["アプリ名", "パッケージ名", "バージョン",
                    "発行元", "インストール日", "管理", "場所"]
        present = set()
        for app in apps:
            present |= set(app.keys())
        present.discard("備考")
        return [c for c in priority if c in present] +                [c for c in present if c not in priority]

    def _sort_tree(self, tv, col, reverse):
        """Treeview の列ヘッダークリックでソート"""
        rows = [(tv.set(k, col), k) for k in tv.get_children("")]
        rows.sort(reverse=reverse,
                  key=lambda x: x[0].lower() if x[0] else "")
        for i, (_, k) in enumerate(rows):
            tv.move(k, "", i)
            tag = "odd" if i % 2 == 0 else "even"
            tv.item(k, tags=(tag,))
        tv.heading(col, command=lambda: self._sort_tree(tv, col, not reverse))


def main():
    app = PCSpecApp()
    app.mainloop()


if __name__ == "__main__":
    main()
